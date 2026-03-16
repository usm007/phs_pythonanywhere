from flask import Flask, render_template, request, redirect, url_for, flash, Response, send_from_directory, session, abort, jsonify, g
import os
import csv
import io
import re
import json
import sqlite3
from datetime import datetime
from functools import wraps

from werkzeug.security import check_password_hash, generate_password_hash
from werkzeug.utils import secure_filename


def normalize_dob(raw_dob):
    """Normalize a DOB string to DD/MM/YYYY. Accepts YYYY-MM-DD, DD/MM/YYYY, DD/MM/YY, D/M/YY etc."""
    s = (raw_dob or "").strip()
    if not s:
        return ""
    # YYYY-MM-DD  →  DD/MM/YYYY
    m = re.match(r"^(\d{4})-(\d{2})-(\d{2})$", s)
    if m:
        return f"{m.group(3)}/{m.group(2)}/{m.group(1)}"
    # DD/MM/YYYY (already correct, normalize leading zeros)
    m = re.match(r"^(\d{1,2})/(\d{1,2})/(\d{4})$", s)
    if m:
        return f"{m.group(1).zfill(2)}/{m.group(2).zfill(2)}/{m.group(3)}"
    # DD/MM/YY  →  DD/MM/YYYY (expand 2-digit year: ≤30 → 20xx, else → 19xx)
    m = re.match(r"^(\d{1,2})/(\d{1,2})/(\d{2})$", s)
    if m:
        dd = m.group(1).zfill(2)
        mm = m.group(2).zfill(2)
        yy = int(m.group(3))
        yyyy = f"20{m.group(3)}" if yy <= 30 else f"19{m.group(3)}"
        return f"{dd}/{mm}/{yyyy}"
    # DD-MM-YYYY or DD-MM-YY
    m = re.match(r"^(\d{1,2})-(\d{1,2})-(\d{2,4})$", s)
    if m:
        dd = m.group(1).zfill(2)
        mm = m.group(2).zfill(2)
        yr = m.group(3)
        yyyy = yr if len(yr) == 4 else (f"20{yr}" if int(yr) <= 30 else f"19{yr}")
        return f"{dd}/{mm}/{yyyy}"
    return s  # unrecognized format: store as-is

try:
    from openpyxl import Workbook, load_workbook as _load_workbook
except ImportError:
    Workbook = None
    _load_workbook = None

app = Flask(__name__)
app.secret_key = os.environ.get("PHS_SECRET_KEY", "change-me-in-env")

# Gzip compression for all responses
try:
    from flask_compress import Compress
    Compress(app)
except ImportError:
    pass  # Optional dependency — degrades gracefully

# Long-lived cache for versioned static assets
@app.after_request
def add_cache_headers(response):
    if request.path.startswith('/static/'):
        response.cache_control.public = True
        response.cache_control.max_age = 31536000  # 1 year
        response.cache_control.immutable = True
    return response

try:
    import pymysql
    import pymysql.cursors
except ImportError:
    pymysql = None

# TiDB configuration (set via environment variables)
DB_HOST = os.environ.get("TIDB_HOST", "")
DB_USER = os.environ.get("TIDB_USER", "")
DB_PASSWORD = os.environ.get("TIDB_PASSWORD", "")
DB_NAME = os.environ.get("TIDB_DATABASE", "test")
DB_PORT = int(os.environ.get("TIDB_PORT", "4000"))
DB_MODE = os.environ.get("PHS_DB_MODE", "sqlite").lower()
DB_CONNECT_TIMEOUT_SECONDS = int(os.environ.get("PHS_DB_CONNECT_TIMEOUT", "3"))
SEED_SAMPLE_DATA = os.environ.get("PHS_SEED_SAMPLE_DATA", "false").lower() == "true"
BATCH_DELETE_PASSWORD = os.environ.get("PHS_BATCH_DELETE_PASSWORD", "phs")
DEFAULT_ADMIN_PIN = os.environ.get("PHS_ADMIN_PIN", "2026")

if DB_MODE not in {"auto", "tidb", "sqlite"}:
    DB_MODE = "auto"

THIS_FOLDER = os.path.dirname(os.path.abspath(__file__))
LOCAL_DB_FILE = os.path.join(THIS_FOLDER, "school_marks.db")
STATIC_FOLDER = os.path.join(THIS_FOLDER, "static")
UPLOADS_FOLDER = os.path.join(STATIC_FOLDER, "uploads")
SCHOOL_LOGO_FILENAME = "school_logo.png"
EMPTY_CLASS_SUBJECT = "__PHS_EMPTY_CLASS__"

SETTINGS_DEFAULTS = {
    "school_name": "Phulani High School",
    "school_address": "Majuli",
    "school_shortcode": "PHS",
    "exam_name": "Annual Examination 2026",
    "academic_session": "2025-26",
    "portal_locked": "0",
    "school_logo_updated_at": "",
    "hidden_panels": "[]",
    "visitor_count": "0",
}


def setting_bool(value):
    return str(value).strip().lower() in {"1", "true", "yes", "on"}


def _connect_tidb():
    if pymysql is None:
        raise RuntimeError("PyMySQL is not installed")

    if not (DB_HOST and DB_USER and DB_PASSWORD):
        raise RuntimeError("TiDB credentials are not configured")

    return pymysql.connect(
        host=DB_HOST,
        port=DB_PORT,
        user=DB_USER,
        password=DB_PASSWORD,
        database=DB_NAME,
        cursorclass=pymysql.cursors.DictCursor,
        autocommit=False,
        ssl_verify_cert=True,
        ssl_verify_identity=True,
        connect_timeout=DB_CONNECT_TIMEOUT_SECONDS,
        read_timeout=DB_CONNECT_TIMEOUT_SECONDS,
        write_timeout=DB_CONNECT_TIMEOUT_SECONDS,
    )


def _connect_sqlite():
    conn = sqlite3.connect(LOCAL_DB_FILE)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    return conn


def get_db_connection():
    if DB_MODE == "sqlite":
        return _connect_sqlite()
    if DB_MODE == "tidb":
        return _connect_tidb()

    # auto mode: prefer cloud, gracefully fall back offline.
    try:
        return _connect_tidb()
    except Exception:
        return _connect_sqlite()


def _is_sqlite_connection(conn):
    return isinstance(conn, sqlite3.Connection)


def _adapt_query_for_backend(conn, query):
    if not _is_sqlite_connection(conn):
        return query
    return query.replace("%s", "?").replace("AS UNSIGNED", "AS INTEGER")


def fetch_all(conn, query, params=()):
    cursor = conn.cursor()
    try:
        cursor.execute(_adapt_query_for_backend(conn, query), params)
        return cursor.fetchall()
    finally:
        cursor.close()


def fetch_one(conn, query, params=()):
    cursor = conn.cursor()
    try:
        cursor.execute(_adapt_query_for_backend(conn, query), params)
        return cursor.fetchone()
    finally:
        cursor.close()


def execute_stmt(conn, query, params=()):
    cursor = conn.cursor()
    try:
        cursor.execute(_adapt_query_for_backend(conn, query), params)
    finally:
        cursor.close()


def executemany_stmt(conn, query, rows):
    cursor = conn.cursor()
    try:
        cursor.executemany(_adapt_query_for_backend(conn, query), rows)
    finally:
        cursor.close()


def log_change(
    conn,
    action,
    entity_type,
    details,
    class_name=None,
    subject=None,
    affected_count=1,
):
    execute_stmt(
        conn,
        """
        INSERT INTO change_logs (
            action, entity_type, class_name, subject, details, affected_count, created_at
        ) VALUES (%s, %s, %s, %s, %s, %s, %s)
        """,
        (
            action,
            entity_type,
            class_name,
            subject,
            details,
            affected_count,
            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        ),
    )


def get_recent_logs(limit=20, conn=None):
    _close = conn is None
    if _close:
        conn = get_db_connection()
    try:
        rows = fetch_all(
            conn,
            """
            SELECT id, action, entity_type, class_name, subject, details, affected_count, created_at
            FROM change_logs
            ORDER BY id DESC
            LIMIT %s
            """,
            (limit,),
        )
        return rows
    finally:
        if _close:
            conn.close()

SUBJECTS_DICT = {
    "Class 6": [
        "Assamese",
        "English",
        "General Science",
        "General Mathematics",
        "Social Science",
        "Hindi",
    ],
    "Class 7": [
        "Assamese",
        "English",
        "General Science",
        "General Mathematics",
        "Social Science",
        "Hindi",
    ],
    "Class 8(A)": [
        "Assamese",
        "English",
        "General Science",
        "General Mathematics",
        "Social Science",
        "Hindi",
    ],
    "Class 8(B)": [
        "Assamese",
        "English",
        "General Science",
        "General Mathematics",
        "Social Science",
        "Hindi",
    ],
    "Class 9(A)": [
        "Assamese",
        "English",
        "General Science",
        "General Mathematics",
        "Social Science",
        "Elective: Hindi",
        "Elective: History",
        "Elective: Agriculture & Horticulture NSQF",
        "Elective: Healthcare NSQF",
    ],
    "Class 9(B)": [
        "Assamese",
        "English",
        "General Science",
        "General Mathematics",
        "Social Science",
        "Elective: Hindi",
        "Elective: History",
        "Elective: Agriculture & Horticulture NSQF",
        "Elective: Healthcare NSQF",
    ],
}

CLASS_DISPLAY_MAP = {
    "Class 6": "Class VI",
    "Class 7": "Class VII",
    "Class 8(A)": "Class VIII(A)",
    "Class 8(B)": "Class VIII(B)",
    "Class 9(A)": "Class IX(A)",
    "Class 9(B)": "Class IX(B)",
}

PROMOTION_MAP_DEFAULT = {
    "Class 6": "Class VII",
    "Class 7": "Class VIII",
    "Class 8(A)": "Class IX(A)",
    "Class 8(B)": "Class IX(B)",
    "Class 9(A)": "Class X",
    "Class 9(B)": "Class X",
}


def get_promotion_map():
    """Return the promotion map, reading from portal_settings if configured."""
    try:
        settings = get_portal_settings()
        raw = settings.get("promotion_map_json", "")
        if raw:
            loaded = json.loads(raw)
            if isinstance(loaded, dict):
                return loaded
    except Exception:
        pass
    return dict(PROMOTION_MAP_DEFAULT)


def class_label(class_name):
    return CLASS_DISPLAY_MAP.get(class_name, class_name)


def _normalize_class_token(value):
    return re.sub(r"[^a-z0-9]+", "", (value or "").strip().lower())


CLASS_IMPORT_ALIASES = {
    # Canonical names
    _normalize_class_token("Class 6"): "Class 6",
    _normalize_class_token("Class 7"): "Class 7",
    _normalize_class_token("Class 8(A)"): "Class 8(A)",
    _normalize_class_token("Class 8(B)"): "Class 8(B)",
    _normalize_class_token("Class 9(A)"): "Class 9(A)",
    _normalize_class_token("Class 9(B)"): "Class 9(B)",
    # Shorthand styles accepted in CSV
    "6": "Class 6",
    "6a": "Class 6",
    "6b": "Class 6",
    "class6": "Class 6",
    "7": "Class 7",
    "7a": "Class 7",
    "7b": "Class 7",
    "class7": "Class 7",
    "8a": "Class 8(A)",
    "class8a": "Class 8(A)",
    "8b": "Class 8(B)",
    "class8b": "Class 8(B)",
    "9a": "Class 9(A)",
    "class9a": "Class 9(A)",
    "9b": "Class 9(B)",
    "class9b": "Class 9(B)",
}


def canonicalize_class_name(value):
    token = _normalize_class_token(value)
    direct = CLASS_IMPORT_ALIASES.get(token)
    if direct:
        return direct

    for class_name in get_subjects_dict().keys():
        if _normalize_class_token(class_name) == token:
            return class_name
    return None


def get_subjects_dict(conn=None):
    owns_connection = conn is None
    if owns_connection:
        conn = get_db_connection()
    try:
        rows = fetch_all(
            conn,
            """
            SELECT class_name, subject
            FROM class_subjects
            ORDER BY class_name, sort_order, subject
            """,
        )
        if not rows:
            return {k: list(v) for k, v in SUBJECTS_DICT.items()}

        dynamic = {}
        for row in rows:
            class_name = row["class_name"]
            subject = row["subject"]
            dynamic.setdefault(class_name, [])
            if subject != EMPTY_CLASS_SUBJECT:
                dynamic[class_name].append(subject)
        return dynamic
    finally:
        if owns_connection:
            conn.close()


def upsert_class_subject(conn, class_name, subject, sort_order):
    if _is_sqlite_connection(conn):
        execute_stmt(
            conn,
            """
            INSERT INTO class_subjects (class_name, subject, sort_order)
            VALUES (%s, %s, %s)
            ON CONFLICT(class_name, subject)
            DO UPDATE SET sort_order=excluded.sort_order
            """,
            (class_name, subject, sort_order),
        )
        return

    execute_stmt(
        conn,
        """
        INSERT INTO class_subjects (class_name, subject, sort_order)
        VALUES (%s, %s, %s)
        ON DUPLICATE KEY UPDATE sort_order=VALUES(sort_order)
        """,
        (class_name, subject, sort_order),
    )


def get_next_subject_order(conn, class_name):
    row = fetch_one(
        conn,
        "SELECT COALESCE(MAX(sort_order), 0) AS m FROM class_subjects WHERE class_name = %s",
        (class_name,),
    )
    return int(row["m"] or 0) + 1


def ensure_class_subject_seed(conn):
    row = fetch_one(conn, "SELECT COUNT(*) AS c FROM class_subjects")
    if row and row["c"]:
        return

    for class_name, subjects in SUBJECTS_DICT.items():
        for idx, subject in enumerate(subjects, start=1):
            upsert_class_subject(conn, class_name, subject, idx)


def get_csrf_token():
    token = session.get("_csrf_token")
    if not token:
        token = os.urandom(24).hex()
        session["_csrf_token"] = token
    return token




@app.before_request
def validate_post_requests():
    if request.method != "POST":
        return

    csrf_token = request.form.get("csrf_token") or request.headers.get("X-CSRF-Token")
    if not csrf_token or csrf_token != session.get("_csrf_token"):
        abort(400, description="Invalid CSRF token")


@app.context_processor
def inject_class_label_helper():
    portal_settings = get_portal_settings()
    try:
        hidden_panels = json.loads(portal_settings.get("hidden_panels", "[]"))
        if not isinstance(hidden_panels, list):
            hidden_panels = []
    except Exception:
        hidden_panels = []
    return {
        "class_label": class_label,
        "csrf_token": get_csrf_token,
        "portal_settings": portal_settings,
        "school_name": portal_settings["school_name"],
        "school_address": portal_settings["school_address"],
        "school_shortcode": portal_settings["school_shortcode"],
        "exam_name": portal_settings["exam_name"],
        "academic_session": portal_settings["academic_session"],
        "school_logo_url": get_school_logo_url(
            portal_settings.get("school_logo_updated_at", "")
        ),
        "portal_locked": setting_bool(portal_settings.get("portal_locked", "0")),
        "admin_unlocked": bool(session.get("admin_unlocked")),
        "hidden_panels": hidden_panels,
    }


def get_school_logo_path():
    return os.path.join(UPLOADS_FOLDER, SCHOOL_LOGO_FILENAME)


def get_school_logo_url(updated_marker=""):
    logo_path = get_school_logo_path()
    if not os.path.exists(logo_path):
        return None
    version = updated_marker or str(int(os.path.getmtime(logo_path)))
    return url_for("static", filename=f"uploads/{SCHOOL_LOGO_FILENAME}", v=version)


def get_portal_settings(conn=None):
    # P1: cache per-request so the context processor doesn't open a new DB conn every call
    if conn is None:
        cached = getattr(g, '_portal_settings', None)
        if cached is not None:
            return cached
    owns_connection = conn is None
    if owns_connection:
        conn = get_db_connection()
    try:
        rows = fetch_all(conn, "SELECT setting_key, setting_value FROM portal_settings")
        merged = dict(SETTINGS_DEFAULTS)
        for row in rows:
            merged[row["setting_key"]] = row["setting_value"]
        if owns_connection:
            g._portal_settings = merged
        return merged
    finally:
        if owns_connection:
            conn.close()


def is_portal_locked(conn=None):
    return setting_bool(get_portal_settings(conn).get("portal_locked", "0"))


def set_setting(conn, key, value):
    execute_stmt(
        conn,
        """
        INSERT INTO portal_settings (setting_key, setting_value)
        VALUES (%s, %s)
        ON DUPLICATE KEY UPDATE setting_value = VALUES(setting_value)
        """,
        (key, value),
    )


def set_setting_sqlite_safe(conn, key, value):
    if _is_sqlite_connection(conn):
        execute_stmt(
            conn,
            """
            INSERT INTO portal_settings (setting_key, setting_value)
            VALUES (%s, %s)
            ON CONFLICT(setting_key) DO UPDATE SET setting_value=excluded.setting_value
            """,
            (key, value),
        )
        return
    set_setting(conn, key, value)


def ensure_setting_defaults(conn):
    existing_rows = fetch_all(conn, "SELECT setting_key FROM portal_settings")
    existing_keys = {row["setting_key"] for row in existing_rows}
    for key, value in SETTINGS_DEFAULTS.items():
        if key not in existing_keys:
            set_setting_sqlite_safe(conn, key, value)


def get_admin_pin_hash(conn=None):
    owns_connection = conn is None
    if owns_connection:
        conn = get_db_connection()
    try:
        row = fetch_one(
            conn,
            "SELECT setting_value FROM portal_settings WHERE setting_key = %s",
            ("admin_pin_hash",),
        )
        return row["setting_value"] if row else None
    finally:
        if owns_connection:
            conn.close()


def set_admin_pin_hash(conn, pin_hash):
    set_setting_sqlite_safe(conn, "admin_pin_hash", pin_hash)


def increment_visitor_count(conn):
    row = fetch_one(
        conn,
        "SELECT setting_value FROM portal_settings WHERE setting_key = %s",
        ("visitor_count",),
    )
    raw_value = row["setting_value"] if row else "0"
    try:
        current_count = int(str(raw_value).strip() or "0")
    except ValueError:
        current_count = 0
    next_count = current_count + 1
    set_setting_sqlite_safe(conn, "visitor_count", str(next_count))
    return next_count


def require_admin_view():
    if session.get("admin_unlocked"):
        return None
    flash("Admin access is locked. Enter Master PIN.", "warning")
    return redirect(url_for("admin_lock"))


def admin_required(fn):
    @wraps(fn)
    def wrapped(*args, **kwargs):
        gate = require_admin_view()
        if gate is not None:
            return gate
        return fn(*args, **kwargs)

    return wrapped


def block_if_locked():
    if is_portal_locked():
        flash("Portal is locked. Editing and save actions are disabled.", "warning")
        return True
    return False


def _migrate_dob_to_4digit_year(conn):
    """One-time migration: expand any DD/MM/YY (2-digit year) DOB values to DD/MM/YYYY."""
    import re as _re
    try:
        rows = fetch_all(conn, "SELECT id, dob FROM students WHERE dob IS NOT NULL AND dob != %s", ("",))
    except Exception:
        return
    updates = []
    for row in rows:
        s = (row["dob"] or "").strip()
        m = _re.match(r"^(\d{2})/(\d{2})/(\d{2})$", s)
        if m:
            yy = int(m.group(3))
            yyyy = ("20" if yy <= 30 else "19") + m.group(3)
            updates.append((f"{m.group(1)}/{m.group(2)}/{yyyy}", row["id"]))
    if updates:
        for new_dob, sid in updates:
            try:
                execute_stmt(conn, "UPDATE students SET dob = %s WHERE id = %s", (new_dob, sid))
            except Exception:
                pass
        try:
            conn.commit()
        except Exception:
            pass


def init_db():
    conn = get_db_connection()
    try:
        if _is_sqlite_connection(conn):
            execute_stmt(
                conn,
                """
                CREATE TABLE IF NOT EXISTS students (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    roll_no TEXT NOT NULL,
                    name TEXT NOT NULL,
                    class_name TEXT NOT NULL,
                    dob TEXT NOT NULL DEFAULT '',
                    father_name TEXT NOT NULL DEFAULT '',
                    mother_name TEXT NOT NULL DEFAULT ''
                )
                """,
            )
            execute_stmt(
                conn,
                """
                CREATE TABLE IF NOT EXISTS marks (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    student_id INTEGER NOT NULL,
                    subject TEXT NOT NULL,
                    marks_obtained INTEGER NOT NULL,
                    UNIQUE(student_id, subject),
                    FOREIGN KEY(student_id) REFERENCES students(id) ON DELETE CASCADE
                )
                """,
            )
            execute_stmt(
                conn,
                """
                CREATE TABLE IF NOT EXISTS change_logs (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    action TEXT NOT NULL,
                    entity_type TEXT NOT NULL,
                    class_name TEXT,
                    subject TEXT,
                    details TEXT NOT NULL,
                    affected_count INTEGER NOT NULL DEFAULT 1,
                    created_at TEXT NOT NULL
                )
                """,
            )
            execute_stmt(
                conn,
                """
                CREATE TABLE IF NOT EXISTS portal_settings (
                    setting_key TEXT PRIMARY KEY,
                    setting_value TEXT NOT NULL
                )
                """,
            )
            execute_stmt(
                conn,
                """
                CREATE TABLE IF NOT EXISTS class_subjects (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    class_name TEXT NOT NULL,
                    subject TEXT NOT NULL,
                    sort_order INTEGER NOT NULL DEFAULT 0,
                    UNIQUE(class_name, subject)
                )
                """,
            )
            execute_stmt(
                conn,
                """
                CREATE TABLE IF NOT EXISTS class_exams (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    class_name TEXT NOT NULL,
                    exam_name TEXT NOT NULL,
                    sort_order INTEGER NOT NULL DEFAULT 0,
                    UNIQUE(class_name, exam_name)
                )
                """,
            )
            execute_stmt(
                conn,
                """
                CREATE TABLE IF NOT EXISTS exam_subject_maxmarks (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    class_name TEXT NOT NULL,
                    exam_name TEXT NOT NULL,
                    subject TEXT NOT NULL,
                    max_marks INTEGER NOT NULL DEFAULT 100,
                    UNIQUE(class_name, exam_name, subject)
                )
                """,
            )
        else:
            execute_stmt(
                conn,
                """
                CREATE TABLE IF NOT EXISTS students (
                    id BIGINT PRIMARY KEY AUTO_INCREMENT,
                    roll_no VARCHAR(32) NOT NULL,
                    name VARCHAR(255) NOT NULL,
                    class_name VARCHAR(32) NOT NULL,
                    dob VARCHAR(32) NOT NULL DEFAULT '',
                    father_name VARCHAR(255) NOT NULL DEFAULT '',
                    mother_name VARCHAR(255) NOT NULL DEFAULT ''
                )
                """,
            )
            execute_stmt(
                conn,
                """
                CREATE TABLE IF NOT EXISTS marks (
                    id BIGINT PRIMARY KEY AUTO_INCREMENT,
                    student_id BIGINT NOT NULL,
                    subject VARCHAR(255) NOT NULL,
                    marks_obtained INT NOT NULL,
                    UNIQUE KEY uq_student_subject (student_id, subject),
                    CONSTRAINT fk_marks_student
                        FOREIGN KEY (student_id) REFERENCES students(id) ON DELETE CASCADE
                )
                """,
            )
            execute_stmt(
                conn,
                """
                CREATE TABLE IF NOT EXISTS change_logs (
                    id BIGINT PRIMARY KEY AUTO_INCREMENT,
                    action VARCHAR(64) NOT NULL,
                    entity_type VARCHAR(64) NOT NULL,
                    class_name VARCHAR(64) NULL,
                    subject VARCHAR(255) NULL,
                    details VARCHAR(1000) NOT NULL,
                    affected_count INT NOT NULL DEFAULT 1,
                    created_at VARCHAR(32) NOT NULL
                )
                """,
            )
            execute_stmt(
                conn,
                """
                CREATE TABLE IF NOT EXISTS portal_settings (
                    setting_key VARCHAR(64) PRIMARY KEY,
                    setting_value VARCHAR(1000) NOT NULL
                )
                """,
            )
            execute_stmt(
                conn,
                """
                CREATE TABLE IF NOT EXISTS class_subjects (
                    id BIGINT PRIMARY KEY AUTO_INCREMENT,
                    class_name VARCHAR(64) NOT NULL,
                    subject VARCHAR(255) NOT NULL,
                    sort_order INT NOT NULL DEFAULT 0,
                    UNIQUE KEY uq_class_subject (class_name, subject)
                )
                """,
            )
            execute_stmt(
                conn,
                """
                CREATE TABLE IF NOT EXISTS class_exams (
                    id BIGINT PRIMARY KEY AUTO_INCREMENT,
                    class_name VARCHAR(64) NOT NULL,
                    exam_name VARCHAR(128) NOT NULL,
                    sort_order INT NOT NULL DEFAULT 0,
                    UNIQUE KEY uq_class_exam (class_name, exam_name)
                )
                """,
            )
            execute_stmt(
                conn,
                """
                CREATE TABLE IF NOT EXISTS exam_subject_maxmarks (
                    id BIGINT PRIMARY KEY AUTO_INCREMENT,
                    class_name VARCHAR(64) NOT NULL,
                    exam_name VARCHAR(128) NOT NULL,
                    subject VARCHAR(128) NOT NULL,
                    max_marks INT NOT NULL DEFAULT 100,
                    UNIQUE KEY uq_exam_sub_max (class_name, exam_name, subject)
                )
                """,
            )

        # Migrate: add new student columns for existing databases
        if _is_sqlite_connection(conn):
            _new_student_cols = [
                ("dob", "TEXT NOT NULL DEFAULT ''"),
                ("father_name", "TEXT NOT NULL DEFAULT ''"),
                ("mother_name", "TEXT NOT NULL DEFAULT ''"),
            ]
        else:
            _new_student_cols = [
                ("dob", "VARCHAR(32) NOT NULL DEFAULT ''"),
                ("father_name", "VARCHAR(255) NOT NULL DEFAULT ''"),
                ("mother_name", "VARCHAR(255) NOT NULL DEFAULT ''"),
            ]
        for _col_name, _col_def in _new_student_cols:
            try:
                execute_stmt(conn, f"ALTER TABLE students ADD COLUMN {_col_name} {_col_def}")
            except Exception:
                pass  # Column already exists

        # Migrate: enforce uniqueness on (class_name, roll_no) to prevent duplicate CSV imports.
        # Deduplicate any existing rows first (keep the lowest id per group so marks are preserved).
        if _is_sqlite_connection(conn):
            try:
                execute_stmt(
                    conn,
                    "CREATE UNIQUE INDEX IF NOT EXISTS uq_students_class_roll ON students (class_name, roll_no)",
                )
            except Exception:
                # Existing duplicates are blocking index creation – remove them first.
                try:
                    execute_stmt(
                        conn,
                        "DELETE FROM students WHERE id NOT IN (SELECT MIN(id) FROM students GROUP BY class_name, roll_no)",
                    )
                    execute_stmt(
                        conn,
                        "CREATE UNIQUE INDEX IF NOT EXISTS uq_students_class_roll ON students (class_name, roll_no)",
                    )
                except Exception:
                    pass
        else:
            try:
                execute_stmt(
                    conn,
                    "ALTER TABLE students ADD UNIQUE KEY uq_students_class_roll (class_name, roll_no)",
                )
            except Exception:
                pass  # Key already exists

        # Migrate: add exam_name column to marks and update unique constraint
        if _is_sqlite_connection(conn):
            _marks_cols = fetch_all(conn, "PRAGMA table_info(marks)")
            _marks_col_names = [c["name"] for c in _marks_cols]
            if "exam_name" not in _marks_col_names:
                # Recreate to change unique constraint
                execute_stmt(conn, "PRAGMA foreign_keys = OFF")
                execute_stmt(
                    conn,
                    """
                    CREATE TABLE marks_new (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        student_id INTEGER NOT NULL,
                        subject TEXT NOT NULL,
                        marks_obtained INTEGER NOT NULL,
                        exam_name TEXT NOT NULL DEFAULT '',
                        UNIQUE(student_id, subject, exam_name),
                        FOREIGN KEY(student_id) REFERENCES students(id) ON DELETE CASCADE
                    )
                    """,
                )
                execute_stmt(
                    conn,
                    "INSERT INTO marks_new (id, student_id, subject, marks_obtained, exam_name) SELECT id, student_id, subject, marks_obtained, '' FROM marks",
                )
                execute_stmt(conn, "DROP TABLE marks")
                execute_stmt(conn, "ALTER TABLE marks_new RENAME TO marks")
                execute_stmt(conn, "PRAGMA foreign_keys = ON")
        else:
            try:
                execute_stmt(
                    conn,
                    "ALTER TABLE marks ADD COLUMN exam_name VARCHAR(128) NOT NULL DEFAULT ''",
                )
            except Exception:
                pass  # Already exists
            try:
                execute_stmt(conn, "ALTER TABLE marks DROP INDEX uq_student_subject")
            except Exception:
                pass
            try:
                execute_stmt(
                    conn,
                    "ALTER TABLE marks ADD UNIQUE KEY uq_student_subject_exam (student_id, subject, exam_name)",
                )
            except Exception:
                pass

        ensure_setting_defaults(conn)
        ensure_class_subject_seed(conn)
        _migrate_dob_to_4digit_year(conn)
        if not get_admin_pin_hash(conn):
            set_admin_pin_hash(conn, generate_password_hash(DEFAULT_ADMIN_PIN))

        # Migrate: notice_board table
        if _is_sqlite_connection(conn):
            execute_stmt(conn, """
                CREATE TABLE IF NOT EXISTS notice_board (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    author TEXT NOT NULL,
                    message TEXT NOT NULL,
                    created_at TEXT NOT NULL DEFAULT (datetime('now','localtime'))
                )
            """)
        else:
            execute_stmt(conn, """
                CREATE TABLE IF NOT EXISTS notice_board (
                    id BIGINT PRIMARY KEY AUTO_INCREMENT,
                    author VARCHAR(128) NOT NULL,
                    message TEXT NOT NULL,
                    created_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP
                )
            """)

        count_row = fetch_one(conn, "SELECT COUNT(*) AS c FROM students")
        count = count_row["c"] if count_row else 0
        if count == 0 and SEED_SAMPLE_DATA:
            sample_rows = []
            for class_name in get_subjects_dict(conn).keys():
                for i in range(1, 6):
                    sample_rows.append((str(i), f"Student {i} {class_name}", class_name))
            executemany_stmt(
                conn,
                "INSERT INTO students (roll_no, name, class_name) VALUES (%s, %s, %s)",
                sample_rows,
            )
        conn.commit()
    finally:
        conn.close()


def get_class_results(class_name, exam_name=None):
    subjects_dict = get_subjects_dict()
    subjects = subjects_dict.get(class_name, [])

    _num_exams = 1
    _annual_rows = []
    conn = get_db_connection()
    try:
        students = fetch_all(
            conn,
            "SELECT id, roll_no, name FROM students WHERE class_name = %s ORDER BY CAST(roll_no AS UNSIGNED), roll_no",
            (class_name,),
        )
        if exam_name is not None:
            marks = fetch_all(
                conn,
                """
                SELECT m.student_id, m.subject, m.marks_obtained
                FROM marks m
                JOIN students s ON m.student_id = s.id
                WHERE s.class_name = %s AND m.exam_name = %s
                """,
                (class_name, exam_name),
            )
        else:
            marks = fetch_all(
                conn,
                """
                SELECT m.student_id, m.subject, SUM(m.marks_obtained) AS marks_obtained
                FROM marks m
                JOIN students s ON m.student_id = s.id
                WHERE s.class_name = %s
                GROUP BY m.student_id, m.subject
                """,
                (class_name,),
            )
            _num_exams = len(get_class_exams(class_name, conn=conn)) or 1
            # Fetch Annual Exam marks for per-subject pass/fail (< 30/100 = FAIL)
            # Matches any exam whose name contains the word "Annual"
            _annual_rows = fetch_all(
                conn,
                """
                SELECT m.student_id, m.subject, m.marks_obtained
                FROM marks m
                JOIN students s ON m.student_id = s.id
                WHERE s.class_name = %s AND m.exam_name LIKE %s
                """,
                (class_name, "%Annual%"),
            )
    finally:
        conn.close()

    max_per_sub = 100 * _num_exams
    grand_total_possible = len(subjects) * max_per_sub

    # Use actual max marks from DB so percentage (and PASS/FAIL) is accurate
    if exam_name is None:
        _emm = get_exam_maxmarks(class_name)
        _actual_total = sum(
            mm
            for exam_mm in _emm.values()
            for sub, mm in exam_mm.items()
            if sub in subjects
        )
        if _actual_total > 0:
            grand_total_possible = _actual_total

    m_dict = {}
    for mark_row in marks:
        if mark_row["student_id"] not in m_dict:
            m_dict[mark_row["student_id"]] = {}
        m_dict[mark_row["student_id"]][mark_row["subject"]] = mark_row["marks_obtained"]

    _annual_by_student = {}
    for row in _annual_rows:
        _annual_by_student.setdefault(row["student_id"], {})[row["subject"]] = row["marks_obtained"]

    results = []
    for student in students:
        student_marks = m_dict.get(student["id"], {})
        total = sum(student_marks.values())
        percentage = round((total / grand_total_possible) * 100, 2) if total > 0 and grand_total_possible > 0 else 0

        if not student_marks:
            status = "ABSENT"
        elif _annual_by_student:
            student_annual = _annual_by_student.get(student["id"], {})
            if student_annual:
                # FAIL if any subject scored < 30 out of 100 in Annual Examination
                status = "PASS" if all(m >= 30 for m in student_annual.values()) else "FAIL"
            else:
                status = "ABSENT"
        else:
            # No annual exam data — fall back to overall percentage
            status = "PASS" if percentage >= 30 else "FAIL"

        results.append(
            {
                "id": student["id"],
                "roll_no": student["roll_no"],
                "name": student["name"],
                "marks": student_marks,
                "total": total,
                "total_possible": grand_total_possible,
                "max_per_subject": max_per_sub,
                "percentage": percentage,
                "status": status,
                "promoted_to": get_promotion_map().get(class_name, "") if status == "PASS" else None,
            }
        )

    # Assign ranks: only PASS students ranked by total desc, ties get same rank
    appeared = [r for r in results if r["status"] != "ABSENT"]
    appeared_sorted = sorted(appeared, key=lambda x: x["total"], reverse=True)
    rank = 1
    for i, r in enumerate(appeared_sorted):
        if i > 0 and r["total"] < appeared_sorted[i - 1]["total"]:
            rank = i + 1
        r["rank"] = rank if r["status"] == "PASS" else "—"
    for r in results:
        if "rank" not in r:
            r["rank"] = "—"

    return results, subjects


def calculate_class_stats(results):
    total_students = len(results)
    pass_count = sum(1 for row in results if row["status"] == "PASS")
    fail_count = sum(1 for row in results if row["status"] == "FAIL")
    absent_count = sum(1 for row in results if row["status"] == "ABSENT")
    appeared_count = total_students - absent_count
    pass_rate = round((pass_count / appeared_count) * 100, 2) if appeared_count else 0

    return {
        "total_students": total_students,
        "appeared_count": appeared_count,
        "pass_count": pass_count,
        "fail_count": fail_count,
        "absent_count": absent_count,
        "pass_rate": pass_rate,
    }


def get_class_exams(class_name, conn=None):
    """Returns list of exam names for a class, ordered by sort_order."""
    _close = conn is None
    if _close:
        conn = get_db_connection()
    try:
        rows = fetch_all(
            conn,
            "SELECT exam_name FROM class_exams WHERE class_name = %s ORDER BY sort_order, id",
            (class_name,),
        )
        return [r["exam_name"] for r in rows]
    finally:
        if _close:
            conn.close()


def get_all_class_exams():
    """Returns dict of class_name -> [exam_name, ...] ordered by sort_order."""
    conn = get_db_connection()
    try:
        rows = fetch_all(
            conn,
            "SELECT class_name, exam_name FROM class_exams ORDER BY class_name, sort_order, id",
        )
    finally:
        conn.close()
    result = {}
    for row in rows:
        result.setdefault(row["class_name"], []).append(row["exam_name"])
    return result


def get_per_exam_marks(class_name):
    """Returns {student_id: {exam_name: {subject: marks_obtained}}} for all exams in a class."""
    conn = get_db_connection()
    try:
        rows = fetch_all(
            conn,
            """
            SELECT m.student_id, m.exam_name, m.subject, m.marks_obtained
            FROM marks m
            JOIN students s ON m.student_id = s.id
            WHERE s.class_name = %s
            """,
            (class_name,),
        )
    finally:
        conn.close()
    result = {}
    for row in rows:
        sid = row["student_id"]
        ename = row["exam_name"]
        subj = row["subject"]
        result.setdefault(sid, {}).setdefault(ename, {})[subj] = row["marks_obtained"]
    return result


def get_exam_maxmarks(class_name):
    """Returns {exam_name: {subject: max_marks}} for all exams in a class."""
    conn = get_db_connection()
    try:
        rows = fetch_all(
            conn,
            "SELECT exam_name, subject, max_marks FROM exam_subject_maxmarks WHERE class_name = %s",
            (class_name,),
        )
    finally:
        conn.close()
    result = {}
    for row in rows:
        result.setdefault(row["exam_name"], {})[row["subject"]] = row["max_marks"]
    return result


@app.route("/notice/add", methods=["POST"])
def notice_add():
    author = request.form.get("author", "").strip()[:64]
    message = request.form.get("message", "").strip()[:500]
    if not author:
        author = "Anonymous"
    if message:
        conn = get_db_connection()
        try:
            execute_stmt(conn, "INSERT INTO notice_board (author, message) VALUES (%s, %s)", (author, message))
            log_change(conn, action="notice_add", entity_type="notice", details=f"Notice by {author}", affected_count=1)
            conn.commit()
        finally:
            conn.close()
    return redirect(url_for("index"))


@app.route("/notice/delete/<int:notice_id>", methods=["POST"])
def notice_delete(notice_id):
    if not session.get("admin_unlocked"):
        flash("Admin access required to delete notices.", "warning")
        return redirect(url_for("index"))
    conn = get_db_connection()
    try:
        execute_stmt(conn, "DELETE FROM notice_board WHERE id = %s", (notice_id,))
        log_change(conn, action="notice_delete", entity_type="notice", details=f"Deleted notice id={notice_id}", affected_count=1)
        conn.commit()
    finally:
        conn.close()
    return redirect(url_for("index"))


@app.route("/")
def index():
    subjects_dict = get_subjects_dict()
    class_exams_map = get_all_class_exams()

    # F4+F7: build per-class completion summary for homepage
    conn = get_db_connection()
    try:
        visitor_count = increment_visitor_count(conn)
        conn.commit()

        # ── Bulk queries (replaces N+1 per-class loops) ──────────────────────
        # All student counts per class in one query
        student_counts_rows = fetch_all(conn, "SELECT class_name, COUNT(*) AS c FROM students GROUP BY class_name")
        student_counts = {r["class_name"]: r["c"] for r in student_counts_rows}

        # All marks counts per class in one query
        marks_counts_rows = fetch_all(
            conn,
            """
            SELECT s.class_name, COUNT(m.id) AS c
            FROM students s
            LEFT JOIN marks m ON m.student_id = s.id
            GROUP BY s.class_name
            """,
        )
        marks_counts = {r["class_name"]: r["c"] for r in marks_counts_rows}

        class_summary = []
        for class_name, subjects in subjects_dict.items():
            exams = class_exams_map.get(class_name, [])
            student_count = student_counts.get(class_name, 0)
            marks_entered = marks_counts.get(class_name, 0)
            num_exams = len(exams) if exams else 1
            marks_expected = student_count * len(subjects) * num_exams
            pct_complete = round((marks_entered / marks_expected) * 100) if marks_expected > 0 else 0

            class_summary.append({
                "class_name": class_name,
                "student_count": student_count,
                "marks_entered": marks_entered,
                "marks_expected": marks_expected,
                "pct_complete": pct_complete,
                "exams": exams,
                "subjects": subjects,
            })

        # ── Card metadata ──────────────────────────────────────────────────
        all_subjects = list({s for subs in subjects_dict.values() for s in subs})
        total_students_count = sum(c["student_count"] for c in class_summary)
        total_classes_count = len(class_summary)
        total_subjects_count = len(all_subjects)

        # Subjects with no marks — single GROUP BY query instead of per-subject loop
        subjects_with_marks_rows = fetch_all(conn, "SELECT DISTINCT subject FROM marks")
        subjects_with_marks = {r["subject"] for r in subjects_with_marks_rows}
        subjects_pending = sum(1 for s in all_subjects if s not in subjects_with_marks)

        # Last entered subject
        last_entry_row = fetch_one(conn, "SELECT subject FROM marks ORDER BY id DESC LIMIT 1")
        last_entry_subject = last_entry_row["subject"] if last_entry_row else None

        # Total marks records entered + latest exam — combined in one query
        meta_row = fetch_one(conn, "SELECT COUNT(*) AS total, MAX(id) AS max_id FROM marks")
        total_records = meta_row["total"] if meta_row else 0
        latest_exam_row = fetch_one(conn, "SELECT exam_name FROM marks ORDER BY id DESC LIMIT 1") if total_records else None
        latest_exam_name = latest_exam_row["exam_name"] if latest_exam_row else None

        # Classes where all marks are complete (pct == 100)
        classes_complete = sum(1 for c in class_summary if c["pct_complete"] == 100)

        # Overall pct for progress ring
        total_entered_sum = sum(c["marks_entered"] for c in class_summary)
        total_expected_sum = sum(c["marks_expected"] for c in class_summary)
        overall_pct_val = round((total_entered_sum / total_expected_sum) * 100) if total_expected_sum > 0 else 0
        # ───────────────────────────────────────────────────────────────────

        notices = fetch_all(conn, "SELECT id, author, message, created_at FROM notice_board ORDER BY id DESC LIMIT 30")
        recent_activity = get_recent_logs(5, conn=conn)
    finally:
        conn.close()

    # Auto-lock admin after rendering — admin sees delete buttons on this visit
    # but is locked on the next home screen visit
    from flask import make_response
    response = make_response(render_template(
        "index.html",
        subjects_dict=subjects_dict,
        class_exams_map=class_exams_map,
        class_summary=class_summary,
        notices=notices,
        recent_activity=recent_activity,
        subjects_pending=subjects_pending,
        last_entry_subject=last_entry_subject,
        total_records=total_records,
        latest_exam_name=latest_exam_name,
        classes_complete=classes_complete,
        total_classes_count=total_classes_count,
        total_subjects_count=total_subjects_count,
        total_students_count=total_students_count,
        overall_pct_val=overall_pct_val,
        visitor_count=visitor_count,
    ))
    session.pop("admin_unlocked", None)
    return response


@app.route("/logs")
def view_logs():
    recent_logs = get_recent_logs(300)
    return render_template("logs.html", recent_logs=recent_logs)


@app.route("/progress")
def marks_progress():
    subjects_dict = get_subjects_dict()
    class_exams_map = get_all_class_exams()
    conn = get_db_connection()
    try:
        # Bulk: student counts per class
        sc_rows = fetch_all(conn, "SELECT class_name, COUNT(*) AS c FROM students GROUP BY class_name")
        student_counts = {r["class_name"]: r["c"] for r in sc_rows}

        # Bulk: marks counts per class+exam
        mc_rows = fetch_all(
            conn,
            """
            SELECT s.class_name, m.exam_name, COUNT(m.id) AS c
            FROM students s
            LEFT JOIN marks m ON m.student_id = s.id
            GROUP BY s.class_name, m.exam_name
            """,
        )
        marks_by_class_exam = {}
        for r in mc_rows:
            marks_by_class_exam.setdefault(r["class_name"], {})[r["exam_name"] or ""] = r["c"]

        progress_data = []
        for class_name, subjects in subjects_dict.items():
            exams = class_exams_map.get(class_name, [])
            student_count = student_counts.get(class_name, 0)
            class_marks = marks_by_class_exam.get(class_name, {})

            exam_progress = []
            total_entered = 0
            total_expected = 0
            if exams:
                for exam in exams:
                    entered = class_marks.get(exam, 0)
                    expected = student_count * len(subjects)
                    pct = round((entered / expected) * 100) if expected > 0 else 0
                    exam_progress.append({"exam": exam, "entered": entered, "expected": expected, "pct": pct})
                    total_entered += entered
                    total_expected += expected
            else:
                total_expected = student_count * len(subjects)
                total_entered = sum(class_marks.values())

            total_pct = round((total_entered / total_expected) * 100) if total_expected > 0 else 0
            progress_data.append({
                "class_name": class_name,
                "student_count": student_count,
                "exam_progress": exam_progress,
                "total_entered": total_entered,
                "total_expected": total_expected,
                "total_pct": total_pct,
                "subjects": subjects,
            })
    finally:
        conn.close()

    return render_template("progress.html", progress_data=progress_data)


@app.route("/results_center")
def results_center():
    subjects_dict = get_subjects_dict()
    class_exams_map = get_all_class_exams()  # single DB call for all classes
    class_cards = []
    for class_name in subjects_dict.keys():
        exams = class_exams_map.get(class_name, [])
        exam_results = []
        if exams:
            for ex in exams:
                res, _ = get_class_results(class_name, exam_name=ex)
                exam_results.append({
                    "exam_name": ex,
                    "results": res,
                    "stats": calculate_class_stats(res),
                })
        else:
            res, _ = get_class_results(class_name, exam_name=None)
            exam_results.append({
                "exam_name": None,
                "results": res,
                "stats": calculate_class_stats(res),
            })
        class_cards.append({
            "class_name": class_name,
            "exams": exams,
            "exam_results": exam_results,
        })
    return render_template(
        "results_center.html",
        class_cards=class_cards,
    )


@app.route("/subject_entry", methods=["GET", "POST"])
def subject_entry():
    subjects_dict = get_subjects_dict()
    if request.method == "POST":
        if block_if_locked():
            return redirect(url_for("index"))

        class_name = request.form["class_name"]
        subject = request.form["subject"]
        exam_name = request.form.get("exam_name", "")

        inserted_count = 0
        updated_count = 0

        conn = get_db_connection()
        try:
            # Fetch per-subject max marks from DB (fallback 100)
            mm_row = fetch_one(
                conn,
                "SELECT max_marks FROM exam_subject_maxmarks WHERE class_name=%s AND exam_name=%s AND subject=%s",
                (class_name, exam_name, subject),
            ) if exam_name else None
            max_per_sub = int(mm_row["max_marks"]) if mm_row else 100
            for key, value in request.form.items():
                if key.startswith("mark_") and value.strip() != "":
                    student_id = key.split("_")[1]
                    try:
                        mark = int(value)
                    except ValueError:
                        flash("Marks must be whole numbers only.", "danger")
                        return redirect(
                            url_for("subject_entry", class_name=class_name, subject=subject, exam_name=exam_name)
                        )

                    if mark < 0 or mark > max_per_sub:
                        flash(
                            f"Invalid marks for {subject}. Enter between 0 and {max_per_sub}.",
                            "danger",
                        )
                        return redirect(
                            url_for("subject_entry", class_name=class_name, subject=subject, exam_name=exam_name)
                        )

                    existing = fetch_one(
                        conn,
                        "SELECT id FROM marks WHERE student_id = %s AND subject = %s AND exam_name = %s",
                        (student_id, subject, exam_name),
                    )
                    if existing:
                        execute_stmt(
                            conn,
                            "UPDATE marks SET marks_obtained = %s WHERE student_id = %s AND subject = %s AND exam_name = %s",
                            (mark, student_id, subject, exam_name),
                        )
                        updated_count += 1
                    else:
                        execute_stmt(
                            conn,
                            "INSERT INTO marks (student_id, subject, marks_obtained, exam_name) VALUES (%s, %s, %s, %s)",
                            (student_id, subject, mark, exam_name),
                        )
                        inserted_count += 1
            total_changed = inserted_count + updated_count
            if total_changed:
                log_change(
                    conn,
                    action="save_marks",
                    entity_type="marks",
                    class_name=class_name,
                    subject=subject,
                    details=f"Saved marks for {class_name} ({subject}){' [' + exam_name + ']' if exam_name else ''}: {inserted_count} inserted, {updated_count} updated.",
                    affected_count=total_changed,
                )
            conn.commit()
        finally:
            conn.close()
        flash(f"Marks saved for {class_name} ({subject}){' [' + exam_name + ']' if exam_name else ''}", "success")
        return redirect(url_for("index"))

    class_name = request.args.get("class_name")
    subject = request.args.get("subject")
    if not class_name or not subject:
        # Show picker page instead of redirecting with error
        return render_template(
            "subject_entry.html",
            pick_mode=True,
            subjects_dict=subjects_dict,
            class_exams_map=get_all_class_exams(),
            class_name=None,
            subject=None,
            students=[],
            max_per_subject=100,
            exam_name="",
            exams=[],
        )
    if class_name not in subjects_dict:
        flash("Invalid class selected.", "danger")
        return redirect(url_for("index"))
    if subject not in subjects_dict[class_name]:
        flash("Invalid subject selected.", "danger")
        return redirect(url_for("index"))

    exam_name = request.args.get("exam_name", "")
    exams = get_class_exams(class_name)

    # Auto-select the only available exam
    if not exam_name and len(exams) == 1:
        return redirect(url_for("subject_entry", class_name=class_name, subject=subject, exam_name=exams[0]))

    conn = get_db_connection()
    try:
        students = fetch_all(
            conn,
            """
            SELECT s.id, s.roll_no, s.name, m.marks_obtained
            FROM students s
            LEFT JOIN marks m ON s.id = m.student_id AND m.subject = %s AND m.exam_name = %s
            WHERE s.class_name = %s
            ORDER BY CAST(s.roll_no AS UNSIGNED), s.roll_no
            """,
            (subject, exam_name, class_name),
        )
        # F1: use configured max marks if available
        mm_row = fetch_one(
            conn,
            "SELECT max_marks FROM exam_subject_maxmarks WHERE class_name = %s AND exam_name = %s AND subject = %s",
            (class_name, exam_name, subject),
        ) if exam_name else None
    finally:
        conn.close()

    max_per_subject = int(mm_row["max_marks"]) if mm_row else 100
    subjects_dict = get_subjects_dict()
    return render_template(
        "subject_entry.html",
        class_name=class_name,
        subject=subject,
        students=students,
        max_per_subject=max_per_subject,
        exam_name=exam_name,
        exams=exams,
        subjects_dict=subjects_dict,
        class_exams_map=get_all_class_exams(),
    )


@app.route("/grid_entry", methods=["GET", "POST"])
def grid_entry():
    subjects_dict = get_subjects_dict()
    class_name = (
        request.args.get("class_name")
        if request.method == "GET"
        else request.form["class_name"]
    )

    if not class_name or class_name not in subjects_dict:
        if request.method == "GET":
            # Show picker instead of error redirect
            return render_template("grid_entry.html", pick_mode=True, subjects_dict=subjects_dict,
                                   class_name=None, subjects=[], exam_name="", exams=[], students=[], max_marks_dict={},
                                   class_exams_map=get_all_class_exams())
        flash("Select a valid class.", "warning")
        return redirect(url_for("index"))

    subjects = subjects_dict[class_name]

    exam_name = request.form.get("exam_name", "") if request.method == "POST" else request.args.get("exam_name", "")
    exams = get_class_exams(class_name) if class_name in subjects_dict else []

    # Auto-select the only available exam on GET
    if request.method == "GET" and not exam_name and len(exams) == 1:
        return redirect(url_for("grid_entry", class_name=class_name, exam_name=exams[0]))

    # F1: build per-subject max marks from DB (fall back to 100)
    exam_maxmarks_db = get_exam_maxmarks(class_name)
    max_marks_dict = {sub: exam_maxmarks_db.get(exam_name, {}).get(sub, 100) for sub in subjects} if exam_name else {sub: 100 for sub in subjects}
    max_per_subject = 100  # kept for legacy template fallback

    conn = get_db_connection()
    if request.method == "POST":
        if block_if_locked():
            conn.close()
            return redirect(url_for("index"))

        inserted_count = 0
        updated_count = 0
        try:
            for key, value in request.form.items():
                if key.startswith("mark_") and value.strip() != "":
                    parts = key.split("_", 2)
                    student_id, subject = parts[1], parts[2]
                    try:
                        mark = int(value)
                    except ValueError:
                        flash("Marks must be whole numbers only.", "danger")
                        return redirect(url_for("grid_entry", class_name=class_name, exam_name=exam_name))

                    if subject not in subjects:
                        flash("Invalid subject in grid payload.", "danger")
                        return redirect(url_for("grid_entry", class_name=class_name, exam_name=exam_name))

                    sub_max = max_marks_dict.get(subject, 100)
                    if mark < 0 or mark > sub_max:
                        flash(
                            f"Invalid marks for {subject}. Enter between 0 and {sub_max}.",
                            "danger",
                        )
                        return redirect(url_for("grid_entry", class_name=class_name, exam_name=exam_name))

                    existing = fetch_one(
                        conn,
                        "SELECT id FROM marks WHERE student_id = %s AND subject = %s AND exam_name = %s",
                        (student_id, subject, exam_name),
                    )
                    if existing:
                        execute_stmt(
                            conn,
                            "UPDATE marks SET marks_obtained = %s WHERE student_id = %s AND subject = %s AND exam_name = %s",
                            (mark, student_id, subject, exam_name),
                        )
                        updated_count += 1
                    else:
                        execute_stmt(
                            conn,
                            "INSERT INTO marks (student_id, subject, marks_obtained, exam_name) VALUES (%s, %s, %s, %s)",
                            (student_id, subject, mark, exam_name),
                        )
                        inserted_count += 1
            total_changed = inserted_count + updated_count
            if total_changed:
                log_change(
                    conn,
                    action="save_grid_marks",
                    entity_type="marks",
                    class_name=class_name,
                    details=f"Saved grid marks for {class_name}{' [' + exam_name + ']' if exam_name else ''}: {inserted_count} inserted, {updated_count} updated.",
                    affected_count=total_changed,
                )
            conn.commit()
        finally:
            conn.close()
        flash(f"Master Grid saved for {class_name}{' [' + exam_name + ']' if exam_name else ''}!", "success")
        return redirect(url_for("index"))

    students = fetch_all(
        conn,
        "SELECT id, roll_no, name FROM students WHERE class_name = %s ORDER BY CAST(roll_no AS UNSIGNED), roll_no",
        (class_name,),
    )

    marks_dict = {}
    rows = fetch_all(
        conn,
        """
        SELECT m.student_id, m.subject, m.marks_obtained
        FROM marks m
        JOIN students s ON s.id = m.student_id
        WHERE s.class_name = %s AND m.exam_name = %s
        """,
        (class_name, exam_name),
    )
    for row in rows:
        if row["student_id"] not in marks_dict:
            marks_dict[row["student_id"]] = {}
        marks_dict[row["student_id"]][row["subject"]] = row["marks_obtained"]

    conn.close()
    return render_template(
        "grid_entry.html",
        class_name=class_name,
        subjects=subjects,
        students=students,
        marks_dict=marks_dict,
        max_per_subject=max_per_subject,
        max_marks_dict=max_marks_dict,
        exam_name=exam_name,
        exams=exams,
        subjects_dict=subjects_dict,
        class_exams_map=get_all_class_exams(),
    )


@app.route("/view")
def view_marks():
    classes = list(get_subjects_dict().keys())
    exam_name = request.args.get("exam_name")

    selected_class = request.args.get("class_name")
    if selected_class not in classes:
        selected_class = classes[0] if classes else None

    # P2: only compute results for the selected class, not all classes
    grouped = {}
    for class_name in classes:
        if class_name == selected_class:
            results, subjects = get_class_results(class_name, exam_name=exam_name)
            stats = calculate_class_stats(results)
            marks_entered = sum(len(row["marks"]) for row in results)
            marks_expected = len(results) * len(subjects)
            students_with_marks = sum(1 for row in results if len(row["marks"]) > 0)
            grouped[class_name] = {
                "results": results,
                "subjects": subjects,
                "stats": stats,
                "marks_entered": marks_entered,
                "marks_expected": marks_expected,
                "students_with_marks": students_with_marks,
            }
        else:
            # Lightweight placeholder — sidebar only needs class names
            grouped[class_name] = None

    selected_data = grouped.get(selected_class) or {
        "results": [],
        "subjects": [],
        "stats": {
            "total_students": 0,
            "appeared_count": 0,
            "pass_count": 0,
            "fail_count": 0,
            "absent_count": 0,
            "pass_rate": 0,
        },
        "marks_entered": 0,
        "marks_expected": 0,
        "students_with_marks": 0,
    }

    # F6: per-exam marks for the selected class (used for exam-wise summary columns)
    selected_exams = get_class_exams(selected_class) if selected_class else []
    per_exam_data = {}
    if selected_class and not exam_name and len(selected_exams) > 1:
        for ex in selected_exams:
            ex_results, _ = get_class_results(selected_class, exam_name=ex)
            per_exam_data[ex] = {r["id"]: r for r in ex_results}

    return render_template(
        "view.html",
        grouped_data=grouped,
        classes=classes,
        selected_class=selected_class,
        selected_data=selected_data,
        class_exams_map=get_all_class_exams(),
        selected_exam=exam_name or "",
        selected_exams=selected_exams,
        per_exam_data=per_exam_data,
    )


@app.route("/download_csv/<string:class_name>")
def download_csv(class_name):
    if class_name not in get_subjects_dict():
        flash("Invalid class.", "danger")
        return redirect(url_for("results_center"))

    exam_name = request.args.get("exam_name")
    exams = get_class_exams(class_name)
    stream = io.StringIO()
    writer = csv.writer(stream)

    if exam_name:
        results, subjects = get_class_results(class_name, exam_name=exam_name)
        writer.writerow(["Roll", "Name"] + subjects + ["Total", "%", "Result"])
        for row in results:
            writer.writerow(
                [row["roll_no"], row["name"]]
                + [row["marks"].get(sub, "") for sub in subjects]
                + [row["total"], row["percentage"], row["status"]]
            )
    elif not exams:
        subjects = get_subjects_dict().get(class_name, [])
        results, _ = get_class_results(class_name)
        writer.writerow(["Roll", "Name"] + subjects + ["Total", "%", "Result"])
        for row in results:
            writer.writerow(
                [row["roll_no"], row["name"]]
                + [row["marks"].get(sub, "") for sub in subjects]
                + [row["total"], row["percentage"], row["status"]]
            )
    else:
        results, subjects, other_exams = get_final_result_data(class_name, exams)
        headers = (
            subjects
            + ["Annual Total", "Annual %"]
            + [f"{ex} %" for ex in other_exams]
            + ["Avg %", "Result", "Rank"]
        )
        writer.writerow(["Roll", "Name"] + headers)
        for row in results:
            writer.writerow(
                [row["roll_no"], row["name"]]
                + [row["marks"].get(sub, "") for sub in subjects]
                + [row["total"], row["percentage"]]
                + [row["exam_pcts"].get(ex, "") for ex in other_exams]
                + [row["avg_percentage"], row["status"], row["final_rank"]]
            )

    return Response(
        stream.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": f"attachment; filename={class_name}.csv"},
    )


@app.route("/download_result_portal_csv/<string:class_name>")
def download_result_portal_csv(class_name):
    if class_name not in get_subjects_dict():
        flash("Invalid class.", "danger")
        return redirect(url_for("results_center"))

    exam_name = request.args.get("exam_name")
    exams = get_class_exams(class_name)

    conn = get_db_connection()
    try:
        dob_rows = fetch_all(
            conn,
            "SELECT id, dob FROM students WHERE class_name = %s",
            (class_name,),
        )
    finally:
        conn.close()

    dob_map = {row["id"]: (row["dob"] or "") for row in dob_rows}

    stream = io.StringIO()
    writer = csv.writer(stream)

    if exam_name:
        results, subjects = get_class_results(class_name, exam_name=exam_name)
        writer.writerow(["Roll", "Name", "DOB"] + subjects + ["Total", "%", "Result"])
        for row in results:
            dob = dob_map.get(row["id"], "")
            writer.writerow(
                [row["roll_no"], row["name"], dob]
                + [row["marks"].get(sub, "") for sub in subjects]
                + [row["total"], row["percentage"], row["status"]]
            )
    elif not exams:
        subjects = get_subjects_dict().get(class_name, [])
        results, _ = get_class_results(class_name)
        writer.writerow(["Roll", "Name", "DOB"] + subjects + ["Total", "%", "Result"])
        for row in results:
            dob = dob_map.get(row["id"], "")
            writer.writerow(
                [row["roll_no"], row["name"], dob]
                + [row["marks"].get(sub, "") for sub in subjects]
                + [row["total"], row["percentage"], row["status"]]
            )
    else:
        results, subjects, other_exams = get_final_result_data(class_name, exams)
        headers = (
            subjects
            + ["Annual Total", "Annual %"]
            + [f"{ex} %" for ex in other_exams]
            + ["Avg %", "Result", "Rank"]
        )
        writer.writerow(["Roll", "Name", "DOB"] + headers)
        for row in results:
            dob = dob_map.get(row["id"], "")
            writer.writerow(
                [row["roll_no"], row["name"], dob]
                + [row["marks"].get(sub, "") for sub in subjects]
                + [row["total"], row["percentage"]]
                + [row["exam_pcts"].get(ex, "") for ex in other_exams]
                + [row["avg_percentage"], row["status"], row["final_rank"]]
            )

    safe_name = class_name.replace(" ", "_")
    return Response(
        stream.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": f"attachment; filename=result_portal_{safe_name}.csv"},
    )


@app.route("/download_student_import_sample/single/<string:class_name>")
def download_student_import_sample_single(class_name):
    if class_name not in get_subjects_dict():
        flash("Invalid class for sample download.", "danger")
        return redirect(url_for("manage_students"))

    stream = io.StringIO()
    writer = csv.writer(stream)
    writer.writerow(["roll_no", "name", "dob", "father_name", "mother_name"])
    writer.writerow(["1", "Aman Das", "15/08/10", "Raju Das", "Priya Das"])
    writer.writerow(["2", "Riya Sharma", "23/04/11", "Suresh Sharma", "Geeta Sharma"])
    writer.writerow(["3", "Neel Bora", "07/11/10", "Kamal Bora", "Sita Bora"])

    return Response(
        stream.getvalue(),
        mimetype="text/csv",
        headers={
            "Content-Disposition": f"attachment; filename=student_import_single_{class_name}.csv"
        },
    )


@app.route("/download_student_import_sample/multi")
def download_student_import_sample_multi():
    classes = list(get_subjects_dict().keys())
    stream = io.StringIO()
    writer = csv.writer(stream)
    writer.writerow(["class_name", "roll_no", "name", "dob", "father_name", "mother_name"])

    if classes:
        writer.writerow([classes[0], "1", "Aman Das", "15/08/10", "Raju Das", "Priya Das"])
    if len(classes) > 1:
        writer.writerow([classes[1], "12", "Riya Sharma", "23/04/11", "Suresh Sharma", "Geeta Sharma"])
    if len(classes) > 2:
        writer.writerow([classes[2], "4", "Neel Bora", "07/11/10", "Kamal Bora", "Sita Bora"])

    return Response(
        stream.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": "attachment; filename=student_import_multi.csv"},
    )


def get_final_result_data(class_name, class_exams):
    """Build Final Result data: Annual exam marks + per-exam % for other exams + avg % + rank."""
    annual_exam = next((e for e in class_exams if "annual" in e.lower()), None)
    other_exams = [e for e in class_exams if e != annual_exam]

    if annual_exam:
        results, subjects = get_class_results(class_name, exam_name=annual_exam)
    else:
        results, subjects = get_class_results(class_name)
        other_exams = []

    other_exam_pcts = {}
    for ex in other_exams:
        ex_results, _ = get_class_results(class_name, exam_name=ex)
        other_exam_pcts[ex] = {r["id"]: r["percentage"] for r in ex_results}

    final_results = []
    for r in results:
        row = dict(r)
        exam_pcts = {}
        all_pcts = []
        if r["status"] != "ABSENT":
            all_pcts.append(float(r["percentage"]))
        for ex in other_exams:
            pct = other_exam_pcts[ex].get(r["id"])
            exam_pcts[ex] = pct
            if pct is not None:
                all_pcts.append(float(pct))
        row["exam_pcts"] = exam_pcts
        row["avg_percentage"] = round(sum(all_pcts) / len(all_pcts), 2) if all_pcts else 0
        row["final_rank"] = "—"
        final_results.append(row)

    pass_rows = [r for r in final_results if r["status"] == "PASS"]
    pass_rows.sort(key=lambda x: x["avg_percentage"], reverse=True)
    rank = 1
    for i, r in enumerate(pass_rows):
        if i > 0 and r["avg_percentage"] < pass_rows[i - 1]["avg_percentage"]:
            rank = i + 1
        r["final_rank"] = rank

    return final_results, subjects, other_exams


@app.route("/print_class_ledger/<string:class_name>")
def print_class_ledger(class_name):
    if class_name not in get_subjects_dict():
        flash("Invalid class.", "danger")
        return redirect(url_for("results_center"))

    exam_name = request.args.get("exam_name")
    class_exams = get_class_exams(class_name)

    if not exam_name:
        results, subjects, other_exams = get_final_result_data(class_name, class_exams)
        stats = calculate_class_stats(results)
        settings = get_portal_settings()
        return render_template(
            "print_ledger.html",
            class_name=class_name,
            results=results,
            subjects=subjects,
            stats=stats,
            class_exams=class_exams,
            exam_name="",
            is_final_result=True,
            other_exams=other_exams,
            ledger_cfg_json=settings.get("ledger_layout", "null"),
        )

    results, subjects = get_class_results(class_name, exam_name=exam_name)
    stats = calculate_class_stats(results)
    settings = get_portal_settings()
    return render_template(
        "print_ledger.html",
        class_name=class_name,
        results=results,
        subjects=subjects,
        stats=stats,
        class_exams=class_exams,
        exam_name=exam_name,
        is_final_result=False,
        other_exams=[],
        ledger_cfg_json=settings.get("ledger_layout", "null"),
    )


@app.route("/api/ledger-layout/save", methods=["POST"])
def ledger_layout_save():
    try:
        cfg = request.get_json(force=True, silent=True) or {}
        conn = get_db_connection()
        try:
            set_setting_sqlite_safe(conn, "ledger_layout", json.dumps(cfg))
            conn.commit()
        finally:
            conn.close()
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/report_cards/<string:class_name>")
def report_cards_bulk(class_name):
    if class_name not in get_subjects_dict():
        flash("Invalid class.", "danger")
        return redirect(url_for("results_center"))

    exam_name = request.args.get("exam_name")
    all_class_exams = get_class_exams(class_name)
    class_exams = [exam_name] if exam_name else all_class_exams
    results, subjects = get_class_results(class_name, exam_name=exam_name or None)
    per_exam_marks = get_per_exam_marks(class_name)
    exam_maxmarks = get_exam_maxmarks(class_name)
    return render_template(
        "report_card.html",
        class_name=class_name,
        results=results,
        subjects=subjects,
        class_exams=class_exams,
        per_exam_marks=per_exam_marks,
        exam_maxmarks=exam_maxmarks,
        exam_name=exam_name or "",
    )


@app.route("/report_cards/<string:class_name>/individual")
def report_cards_individual_list(class_name):
    if class_name not in get_subjects_dict():
        flash("Invalid class.", "danger")
        return redirect(url_for("results_center"))

    exam_name = request.args.get("exam_name")
    results, _ = get_class_results(class_name, exam_name=exam_name or None)
    return render_template(
        "individual_report_cards.html",
        class_name=class_name,
        results=results,
        exam_name=exam_name or "",
    )


@app.route("/report_cards/<string:class_name>/individual/<int:student_id>")
def report_card_individual(class_name, student_id):
    if class_name not in get_subjects_dict():
        flash("Invalid class.", "danger")
        return redirect(url_for("results_center"))

    exam_name = request.args.get("exam_name")
    all_class_exams = get_class_exams(class_name)
    class_exams = [exam_name] if exam_name else all_class_exams
    results, subjects = get_class_results(class_name, exam_name=exam_name or None)
    per_exam_marks = get_per_exam_marks(class_name)
    exam_maxmarks = get_exam_maxmarks(class_name)
    selected_student = next((row for row in results if row["id"] == student_id), None)
    if not selected_student:
        flash("Student not found in selected class.", "warning")
        return redirect(url_for("report_cards_individual_list", class_name=class_name))

    return render_template(
        "report_card.html",
        class_name=class_name,
        results=[selected_student],
        subjects=subjects,
        single_mode=True,
        class_exams=class_exams,
        per_exam_marks=per_exam_marks,
        exam_maxmarks=exam_maxmarks,
        exam_name=exam_name or "",
    )


@app.route("/manage_students", methods=["GET", "POST"])
def manage_students():
    valid_tabs = {"students", "classes", "subjects", "exams"}
    setup_tab = request.args.get("setup_tab", "students")
    if setup_tab not in valid_tabs:
        setup_tab = "students"

    subjects_dict = get_subjects_dict()
    conn = get_db_connection()
    if request.method == "POST":
        setup_tab = request.form.get("setup_tab", "students")
        if setup_tab not in valid_tabs:
            setup_tab = "students"

        if block_if_locked():
            conn.close()
            return redirect(url_for("manage_students", class_name=request.form.get("class_name") or None, setup_tab=setup_tab))

        class_name = request.form.get("class_name")
        added_count = 0
        updated_count = 0

        new_roll = request.form.get("new_roll", "").strip()
        new_name = request.form.get("new_name", "").strip().upper()
        new_dob = normalize_dob(request.form.get("new_dob", "").strip())
        new_father_name = request.form.get("new_father_name", "").strip().upper()
        new_mother_name = request.form.get("new_mother_name", "").strip().upper()
        if new_roll or new_name:
            if not class_name:
                conn.close()
                flash("Select class before adding a student.", "warning")
                return redirect(url_for("manage_students", setup_tab=setup_tab))
            if not new_roll or not new_name:
                conn.close()
                flash("Enter both roll and name for new student.", "warning")
                return redirect(url_for("manage_students", class_name=class_name, setup_tab=setup_tab))

            duplicate = fetch_one(
                conn,
                "SELECT id FROM students WHERE class_name = %s AND roll_no = %s",
                (class_name, new_roll),
            )
            if duplicate:
                conn.close()
                flash(f"Roll {new_roll} already exists in {class_name}.", "danger")
                return redirect(url_for("manage_students", class_name=class_name, setup_tab=setup_tab))

            execute_stmt(
                conn,
                "INSERT INTO students (roll_no, name, class_name, dob, father_name, mother_name) VALUES (%s, %s, %s, %s, %s, %s)",
                (new_roll, new_name, class_name, new_dob, new_father_name, new_mother_name),
            )
            added_count += 1

        for key, value in request.form.items():
            if key.startswith("name_"):
                student_id = key.split("_")[1]
                roll_no = request.form.get(f"roll_{student_id}", "").strip()
                name = value.strip().upper()
                dob = normalize_dob(request.form.get(f"dob_{student_id}", "").strip())
                father_name = request.form.get(f"father_{student_id}", "").strip().upper()
                mother_name = request.form.get(f"mother_{student_id}", "").strip().upper()
                if not name or not roll_no:
                    continue

                existing = fetch_one(
                    conn,
                    "SELECT id FROM students WHERE class_name = %s AND roll_no = %s AND id != %s",
                    (class_name, roll_no, student_id),
                )
                if existing:
                    conn.close()
                    flash(f"Roll {roll_no} is duplicated. Use unique rolls in class.", "danger")
                    return redirect(url_for("manage_students", class_name=class_name, setup_tab=setup_tab))

                current_row = fetch_one(
                    conn,
                    "SELECT roll_no, name, dob, father_name, mother_name FROM students WHERE id = %s",
                    (student_id,),
                )
                if not current_row:
                    continue
                if (current_row["roll_no"] == roll_no and current_row["name"] == name
                        and (current_row["dob"] or "") == dob
                        and (current_row["father_name"] or "") == father_name
                        and (current_row["mother_name"] or "") == mother_name):
                    continue

                execute_stmt(
                    conn,
                    "UPDATE students SET name = %s, roll_no = %s, dob = %s, father_name = %s, mother_name = %s WHERE id = %s",
                    (name, roll_no, dob, father_name, mother_name, student_id),
                )
                updated_count += 1

        total_changed = added_count + updated_count
        if total_changed:
            log_change(
                conn,
                action="manage_students_save",
                entity_type="students",
                class_name=class_name,
                details=f"Saved student changes for {class_name}: {added_count} added, {updated_count} updated.",
                affected_count=total_changed,
            )

        conn.commit()
        conn.close()
        flash("✅ Updated!", "success")
        return redirect(url_for("manage_students", class_name=class_name, setup_tab=setup_tab))

    class_name = request.args.get("class_name")
    if class_name:
        students = fetch_all(
            conn,
            "SELECT * FROM students WHERE class_name = %s ORDER BY CAST(roll_no AS UNSIGNED), roll_no",
            (class_name,),
        )
    else:
        students = []
    conn.close()
    return render_template(
        "manage_students.html",
        class_name=class_name,
        setup_tab=setup_tab,
        students=students,
        classes=list(subjects_dict.keys()),
        class_subject_map=subjects_dict,
        class_exams_map=get_all_class_exams(),
    )


@app.route("/manage_structure", methods=["POST"])
def manage_structure():
    valid_tabs = {"students", "classes", "subjects", "exams"}
    setup_tab = request.form.get("setup_tab", "subjects")
    if setup_tab not in valid_tabs:
        setup_tab = "subjects"

    if block_if_locked():
        return redirect(url_for("manage_students", class_name=request.form.get("class_name") or None, setup_tab=setup_tab))

    raw_action = (request.form.get("structure_action") or "").strip()
    action = raw_action
    class_name = (request.form.get("class_name") or "").strip()
    selected_class = request.form.get("selected_class") or class_name or None

    if raw_action in {"rename_class", "rename_subject"}:
        flash("Rename is disabled. Use add/delete only.", "warning")
        return redirect(url_for("manage_students", class_name=selected_class, setup_tab=setup_tab))

    target_label = "class"

    conn = get_db_connection()
    try:
        if action == "add_class":
            new_class_name = (request.form.get("new_class_name") or "").strip()
            first_subject = (request.form.get("first_subject") or "").strip()
            if not new_class_name:
                flash(f"{target_label.capitalize()} name is required.", "warning")
                return redirect(url_for("manage_students", class_name=selected_class, setup_tab=setup_tab))

            existing = fetch_one(
                conn,
                "SELECT id FROM class_subjects WHERE class_name = %s LIMIT 1",
                (new_class_name,),
            )
            if existing:
                flash(f"That {target_label} already exists.", "warning")
                return redirect(url_for("manage_students", class_name=new_class_name, setup_tab=setup_tab))

            initial_subject = first_subject or EMPTY_CLASS_SUBJECT
            initial_sort_order = 1 if first_subject else 0
            upsert_class_subject(conn, new_class_name, initial_subject, initial_sort_order)
            details = f"Added new {target_label} '{new_class_name}'."
            if first_subject:
                details = f"Added new {target_label} '{new_class_name}' with first subject '{first_subject}'."
            log_change(
                conn,
                action=f"add_{target_label}",
                entity_type="admin",
                class_name=new_class_name,
                details=details,
                affected_count=1,
            )
            conn.commit()
            flash(f"Added {target_label}: {new_class_name}", "success")
            return redirect(url_for("manage_students", class_name=new_class_name, setup_tab=setup_tab))

        if action == "delete_class":
            delete_class_name = (request.form.get("delete_class_name") or "").strip()
            if not delete_class_name:
                flash(f"Select a {target_label} to delete.", "warning")
                return redirect(url_for("manage_students", class_name=selected_class, setup_tab=setup_tab))

            student_count_row = fetch_one(
                conn,
                "SELECT COUNT(*) AS c FROM students WHERE class_name = %s",
                (delete_class_name,),
            )
            student_count = student_count_row["c"] if student_count_row else 0

            execute_stmt(conn, "DELETE FROM class_subjects WHERE class_name = %s", (delete_class_name,))
            execute_stmt(conn, "DELETE FROM students WHERE class_name = %s", (delete_class_name,))
            log_change(
                conn,
                action=f"delete_{target_label}",
                entity_type="admin",
                class_name=delete_class_name,
                details=f"Deleted {target_label} '{delete_class_name}' with {student_count} student records.",
                affected_count=student_count,
            )
            conn.commit()
            flash(f"Deleted {target_label}: {delete_class_name}", "success")
            return redirect(url_for("manage_students", setup_tab=setup_tab))

        if action == "add_subject":
            class_name = (request.form.get("class_name") or "").strip()
            new_subject = (request.form.get("new_subject") or "").strip()
            if not class_name or not new_subject:
                flash("Select class and enter subject name.", "warning")
                return redirect(url_for("manage_students", class_name=selected_class, setup_tab=setup_tab))

            existing = fetch_one(
                conn,
                "SELECT id FROM class_subjects WHERE class_name = %s AND subject = %s",
                (class_name, new_subject),
            )
            if existing:
                flash("Subject already exists in this class.", "warning")
                return redirect(url_for("manage_students", class_name=class_name, setup_tab=setup_tab))

            # Remove class anchor row when adding first real subject.
            execute_stmt(
                conn,
                "DELETE FROM class_subjects WHERE class_name = %s AND subject = %s",
                (class_name, EMPTY_CLASS_SUBJECT),
            )
            upsert_class_subject(conn, class_name, new_subject, get_next_subject_order(conn, class_name))
            log_change(
                conn,
                action="add_subject",
                entity_type="admin",
                class_name=class_name,
                subject=new_subject,
                details=f"Added subject '{new_subject}' to {class_name}.",
                affected_count=1,
            )
            conn.commit()
            flash("Subject added.", "success")
            return redirect(url_for("manage_students", class_name=class_name, setup_tab=setup_tab))

        if action == "delete_subject":
            class_name = (request.form.get("class_name") or "").strip()
            delete_subject = (request.form.get("delete_subject") or "").strip()
            if not class_name or not delete_subject:
                flash("Class and subject are required for delete.", "warning")
                return redirect(url_for("manage_students", class_name=selected_class, setup_tab=setup_tab))

            execute_stmt(
                conn,
                "DELETE FROM class_subjects WHERE class_name = %s AND subject = %s",
                (class_name, delete_subject),
            )
            execute_stmt(
                conn,
                """
                DELETE FROM marks
                WHERE subject = %s
                  AND student_id IN (SELECT id FROM students WHERE class_name = %s)
                """,
                (delete_subject, class_name),
            )

            remaining_subjects_row = fetch_one(
                conn,
                "SELECT COUNT(*) AS c FROM class_subjects WHERE class_name = %s AND subject <> %s",
                (class_name, EMPTY_CLASS_SUBJECT),
            )
            remaining_subjects = remaining_subjects_row["c"] if remaining_subjects_row else 0
            if remaining_subjects == 0:
                upsert_class_subject(conn, class_name, EMPTY_CLASS_SUBJECT, 0)

            log_change(
                conn,
                action="delete_subject",
                entity_type="admin",
                class_name=class_name,
                subject=delete_subject,
                details=f"Deleted subject '{delete_subject}' from {class_name} and removed related marks.",
                affected_count=1,
            )
            conn.commit()
            flash("Subject deleted.", "success")
            return redirect(url_for("manage_students", class_name=class_name, setup_tab=setup_tab))

        if action == "add_exam":
            class_name = (request.form.get("class_name") or "").strip()
            new_exam_name = (request.form.get("new_exam_name") or "").strip()
            if not class_name or not new_exam_name:
                flash("Select class and enter exam name.", "warning")
                return redirect(url_for("manage_students", class_name=selected_class, setup_tab=setup_tab))

            existing = fetch_one(
                conn,
                "SELECT id FROM class_exams WHERE class_name = %s AND exam_name = %s",
                (class_name, new_exam_name),
            )
            if existing:
                flash("Exam already exists for this class.", "warning")
                return redirect(url_for("manage_students", class_name=class_name, setup_tab=setup_tab))

            # Get next sort order
            max_order_row = fetch_one(
                conn,
                "SELECT MAX(sort_order) AS mx FROM class_exams WHERE class_name = %s",
                (class_name,),
            )
            next_order = (max_order_row["mx"] or 0) + 1 if max_order_row else 1
            execute_stmt(
                conn,
                "INSERT INTO class_exams (class_name, exam_name, sort_order) VALUES (%s, %s, %s)",
                (class_name, new_exam_name, next_order),
            )
            # Insert per-subject max marks
            subjects_for_class = get_subjects_dict().get(class_name, [])
            for subj in subjects_for_class:
                field_key = f"max_marks_{subj.replace(' ', '_')}"
                try:
                    mm = int(request.form.get(field_key, 100))
                except (ValueError, TypeError):
                    mm = 100
                execute_stmt(
                    conn,
                    "INSERT INTO exam_subject_maxmarks (class_name, exam_name, subject, max_marks) VALUES (%s, %s, %s, %s)",
                    (class_name, new_exam_name, subj, mm),
                )
            log_change(
                conn,
                action="add_exam",
                entity_type="admin",
                class_name=class_name,
                details=f"Added exam '{new_exam_name}' to {class_name}.",
                affected_count=1,
            )
            conn.commit()
            flash(f"Exam '{new_exam_name}' added to {class_name}.", "success")
            return redirect(url_for("manage_students", class_name=class_name, setup_tab=setup_tab))

        if action == "delete_exam":
            class_name = (request.form.get("class_name") or "").strip()
            delete_exam_name = (request.form.get("delete_exam_name") or "").strip()
            if not class_name or not delete_exam_name:
                flash("Class and exam name required.", "warning")
                return redirect(url_for("manage_students", class_name=selected_class, setup_tab=setup_tab))

            execute_stmt(
                conn,
                "DELETE FROM class_exams WHERE class_name = %s AND exam_name = %s",
                (class_name, delete_exam_name),
            )
            execute_stmt(
                conn,
                "DELETE FROM exam_subject_maxmarks WHERE class_name = %s AND exam_name = %s",
                (class_name, delete_exam_name),
            )
            log_change(
                conn,
                action="delete_exam",
                entity_type="admin",
                class_name=class_name,
                details=f"Deleted exam '{delete_exam_name}' from {class_name}.",
                affected_count=1,
            )
            conn.commit()
            flash(f"Exam '{delete_exam_name}' deleted from {class_name}.", "success")
            return redirect(url_for("manage_students", class_name=class_name, setup_tab=setup_tab))

        flash("Unknown structure action.", "warning")
        return redirect(url_for("manage_students", class_name=selected_class, setup_tab=setup_tab))
    finally:
        conn.close()


@app.route("/delete_student/<int:student_id>", methods=["POST"])
def delete_student(student_id):
    setup_tab = request.form.get("setup_tab") or request.args.get("setup_tab") or "students"
    if block_if_locked():
        return redirect(url_for("manage_students", class_name=request.form.get("class_name") or None, setup_tab=setup_tab))

    class_name = request.form.get("class_name") or request.args.get("class_name")
    conn = get_db_connection()
    student = fetch_one(
        conn,
        "SELECT id, roll_no, name, class_name FROM students WHERE id = %s",
        (student_id,),
    )
    if not student:
        conn.close()
        flash("Student not found.", "warning")
        return redirect(url_for("manage_students", class_name=class_name, setup_tab=setup_tab))

    execute_stmt(conn, "DELETE FROM students WHERE id = %s", (student_id,))
    log_change(
        conn,
        action="delete_student",
        entity_type="students",
        class_name=student["class_name"],
        details=f"Deleted student {student['name']} (roll {student['roll_no']}) from {student['class_name']}.",
        affected_count=1,
    )
    conn.commit()
    conn.close()
    flash("Student deleted successfully.", "success")
    return redirect(url_for("manage_students", class_name=class_name, setup_tab=setup_tab))


def _parse_student_csv(file_stream, single_class_name=None):
    """Parse student CSV and return rows plus parse stats.

    Supports two formats:
      A) Single-class (single_class_name provided): columns roll_no,name[,dob,father_name,mother_name]
      B) Multi-class  (single_class_name is None):  columns class_name,roll_no,name[,dob,father_name,mother_name]

    Accepted header aliases (case-insensitive):
      roll_no:      roll_no, roll, rollno, roll_number
      name:         name, student_name
      class:        class_name, class
      dob:          dob, date_of_birth, dateofbirth, birth_date
      father_name:  father_name, father, fathername
      mother_name:  mother_name, mother, mothername

    Returns (rows, stats):
      rows: list[(class_name, roll_no, name, dob, father_name, mother_name)]
      stats: {
        total_rows, skipped_empty, duplicate_rows_in_file
      }
    """

    def _normalize_header(value):
        value = (value or "").strip().lower()
        value = re.sub(r"[^a-z0-9]+", "_", value)
        return value.strip("_")

    def _find_header(fieldnames, aliases):
        for raw in fieldnames:
            if _normalize_header(raw) in aliases:
                return raw
        return None

    def _normalize_dob(raw_dob):
        """Normalize a DOB string to DD/MM/YYYY. Accepts YYYY-MM-DD, DD/MM/YYYY, DD/MM/YY, D/M/YY etc."""
        s = (raw_dob or "").strip()
        if not s:
            return ""
        # YYYY-MM-DD  →  DD/MM/YYYY
        m = re.match(r"^(\d{4})-(\d{2})-(\d{2})$", s)
        if m:
            return f"{m.group(3)}/{m.group(2)}/{m.group(1)}"
        # DD/MM/YYYY (already correct, normalize leading zeros)
        m = re.match(r"^(\d{1,2})/(\d{1,2})/(\d{4})$", s)
        if m:
            dd = m.group(1).zfill(2)
            mm = m.group(2).zfill(2)
            return f"{dd}/{mm}/{m.group(3)}"
        # DD/MM/YY  →  DD/MM/YYYY (expand 2-digit year: ≤30 → 20xx, else → 19xx)
        m = re.match(r"^(\d{1,2})/(\d{1,2})/(\d{2})$", s)
        if m:
            dd = m.group(1).zfill(2)
            mm = m.group(2).zfill(2)
            yy = int(m.group(3))
            yyyy = f"20{m.group(3)}" if yy <= 30 else f"19{m.group(3)}"
            return f"{dd}/{mm}/{yyyy}"
        # DD-MM-YYYY or DD-MM-YY
        m = re.match(r"^(\d{1,2})-(\d{1,2})-(\d{2,4})$", s)
        if m:
            dd = m.group(1).zfill(2)
            mm = m.group(2).zfill(2)
            yr = m.group(3)
            yyyy = yr if len(yr) == 4 else (f"20{yr}" if int(yr) <= 30 else f"19{yr}")
            return f"{dd}/{mm}/{yyyy}"
        return s  # unrecognized format: store as-is

    # PythonAnywhere/Werkzeug may provide upload streams that don't implement
    # the full file API expected by TextIOWrapper (e.g. readable()).
    raw = file_stream.read()
    if isinstance(raw, bytes):
        decoded = raw.decode("utf-8-sig", errors="replace")
    else:
        decoded = str(raw)

    reader = csv.DictReader(io.StringIO(decoded))
    fieldnames = reader.fieldnames or []

    if not fieldnames:
        raise ValueError("CSV has no header row.")

    roll_key = _find_header(fieldnames, {"roll_no", "roll", "rollno", "roll_number"})
    name_key = _find_header(fieldnames, {"name", "student_name"})
    class_key = _find_header(fieldnames, {"class_name", "class"})
    dob_key = _find_header(fieldnames, {"dob", "date_of_birth", "dateofbirth", "birth_date"})
    father_key = _find_header(fieldnames, {"father_name", "father", "fathername"})
    mother_key = _find_header(fieldnames, {"mother_name", "mother", "mothername"})

    required = ["roll_no", "name"] if single_class_name is not None else ["class_name", "roll_no", "name"]
    missing = []
    if not roll_key:
        missing.append("roll_no")
    if not name_key:
        missing.append("name")
    if single_class_name is None and not class_key:
        missing.append("class_name")
    if missing:
        raise ValueError(f"Missing required header(s): {', '.join(missing)}")

    deduped = {}
    skipped_empty = 0
    duplicate_rows_in_file = 0
    total_rows = 0

    for raw_row in reader:
        total_rows += 1

        row = {k: (v.strip() if isinstance(v, str) else (v or "")) for k, v in raw_row.items()}

        if single_class_name is not None:
            roll_no = row.get(roll_key, "")
            name = row.get(name_key, "")
            class_name = single_class_name
        else:
            class_name = row.get(class_key, "")
            roll_no = row.get(roll_key, "")
            name = row.get(name_key, "")

        dob = _normalize_dob(row.get(dob_key, "") if dob_key else "")
        father_name = row.get(father_key, "") if father_key else ""
        mother_name = row.get(mother_key, "") if mother_key else ""

        if not roll_no or not name or not class_name:
            skipped_empty += 1
            continue

        key = (class_name, roll_no)
        if key in deduped:
            duplicate_rows_in_file += 1
        deduped[key] = (name, dob, father_name, mother_name)

    rows = [(cls, roll, name, dob, father, mother) for (cls, roll), (name, dob, father, mother) in deduped.items()]
    stats = {
        "total_rows": total_rows,
        "skipped_empty": skipped_empty,
        "duplicate_rows_in_file": duplicate_rows_in_file,
    }

    return rows, stats


@app.route("/import_students", methods=["POST"])
def import_students():
    setup_tab = request.form.get("setup_tab") or "students"
    if block_if_locked():
        return redirect(url_for("manage_students", class_name=request.form.get("current_class_name") or None, setup_tab=setup_tab))

    class_name = request.form.get("import_class_name", "").strip() or None
    file = request.files.get("csv_file")
    redirect_class = class_name or request.form.get("current_class_name", "")

    if not file or file.filename == "":
        flash("No CSV file selected.", "warning")
        return redirect(url_for("manage_students", class_name=redirect_class or None, setup_tab=setup_tab))

    filename = (file.filename or "").lower()
    if not filename.endswith(".csv"):
        flash("Please upload a .csv file exported from Excel.", "warning")
        return redirect(url_for("manage_students", class_name=redirect_class or None, setup_tab=setup_tab))

    try:
        rows, parse_stats = _parse_student_csv(file.stream, single_class_name=class_name)
    except Exception as exc:
        flash(f"Failed to read CSV: {exc}", "danger")
        return redirect(url_for("manage_students", class_name=redirect_class or None, setup_tab=setup_tab))

    if not rows:
        skipped_empty = parse_stats.get("skipped_empty", 0)
        duplicates_in_file = parse_stats.get("duplicate_rows_in_file", 0)
        flash(
            (
                "No valid rows found in CSV. "
                f"Skipped empty rows: {skipped_empty}; duplicate rows in file: {duplicates_in_file}."
            ),
            "warning",
        )
        return redirect(url_for("manage_students", class_name=redirect_class or None, setup_tab=setup_tab))

    # Validate and normalize class names against accepted aliases.
    filtered_rows = []
    skipped_invalid_class = 0
    for cls, roll, name, dob, father_name, mother_name in rows:
        canonical_cls = canonicalize_class_name(cls)
        if not canonical_cls:
            skipped_invalid_class += 1
        else:
            filtered_rows.append((canonical_cls, roll, name, dob, father_name, mother_name))
    rows = filtered_rows

    if not rows:
        flash(
            "No importable rows after class validation. "
            f"Invalid class rows: {skipped_invalid_class}.",
            "warning",
        )
        return redirect(url_for("manage_students", class_name=redirect_class or None, setup_tab=setup_tab))

    inserted = 0
    updated = 0
    unchanged = 0
    conn = get_db_connection()
    try:
        for cls, roll, name, dob, father_name, mother_name in rows:
            existing = fetch_one(
                conn,
                "SELECT id, name, dob, father_name, mother_name FROM students WHERE class_name = %s AND roll_no = %s",
                (cls, roll),
            )
            if existing:
                if (existing["name"] != name
                        or (existing["dob"] or "") != dob
                        or (existing["father_name"] or "") != father_name
                        or (existing["mother_name"] or "") != mother_name):
                    execute_stmt(
                        conn,
                        "UPDATE students SET name = %s, dob = %s, father_name = %s, mother_name = %s WHERE id = %s",
                        (name, dob, father_name, mother_name, existing["id"]),
                    )
                    updated += 1
                else:
                    unchanged += 1
            else:
                _insert_sql = (
                    "INSERT OR IGNORE INTO students (roll_no, name, class_name, dob, father_name, mother_name) VALUES (%s, %s, %s, %s, %s, %s)"
                    if _is_sqlite_connection(conn)
                    else "INSERT IGNORE INTO students (roll_no, name, class_name, dob, father_name, mother_name) VALUES (%s, %s, %s, %s, %s, %s)"
                )
                execute_stmt(conn, _insert_sql, (roll, name, cls, dob, father_name, mother_name))
                inserted += 1
        total_changed = inserted + updated
        scope = class_name or "multiple classes"
        log_change(
            conn,
            action="import_students_csv",
            entity_type="students",
            class_name=class_name,
            details=(
                f"CSV import for {scope}: {inserted} inserted, {updated} updated, "
                f"{unchanged} unchanged, {parse_stats.get('duplicate_rows_in_file', 0)} duplicate rows merged, "
                f"{parse_stats.get('skipped_empty', 0)} skipped empty, {skipped_invalid_class} invalid class."
            ),
            affected_count=total_changed,
        )
        conn.commit()
    except Exception as exc:
        conn.rollback()
        flash(f"Import failed: {exc}", "danger")
        return redirect(url_for("manage_students", class_name=redirect_class or None, setup_tab=setup_tab))
    finally:
        conn.close()

    parts = []
    if inserted:
        parts.append(f"{inserted} imported")
    if updated:
        parts.append(f"{updated} updated")
    if unchanged:
        parts.append(f"{unchanged} unchanged")
    skipped_empty = parse_stats.get("skipped_empty", 0)
    duplicates_in_file = parse_stats.get("duplicate_rows_in_file", 0)
    if duplicates_in_file:
        parts.append(f"{duplicates_in_file} duplicate row(s) merged")
    if skipped_empty:
        parts.append(f"{skipped_empty} empty row(s) skipped")
    if skipped_invalid_class:
        parts.append(f"{skipped_invalid_class} invalid class row(s) skipped")
    flash(f"✅ CSV import complete: {', '.join(parts)}.", "success")
    return redirect(url_for("manage_students", class_name=redirect_class or None, setup_tab=setup_tab))


@app.route("/batch_delete_students", methods=["POST"])
def batch_delete_students():
    setup_tab = request.form.get("setup_tab") or "students"
    if block_if_locked():
        return redirect(url_for("manage_students", class_name=request.form.get("class_name") or None, setup_tab=setup_tab))

    class_name = request.form.get("class_name", "")
    password = request.form.get("batch_delete_password", "")

    if password != BATCH_DELETE_PASSWORD:
        flash("Incorrect password. Batch delete cancelled.", "danger")
        return redirect(url_for("manage_students", class_name=class_name or None, setup_tab=setup_tab))

    student_ids = request.form.getlist("delete_ids")
    if not student_ids:
        flash("No students selected for deletion.", "warning")
        return redirect(url_for("manage_students", class_name=class_name or None, setup_tab=setup_tab))

    conn = get_db_connection()
    try:
        deleted = 0
        for sid in student_ids:
            try:
                sid_int = int(sid)
            except ValueError:
                continue
            existing = fetch_one(
                conn, "SELECT id FROM students WHERE id = %s", (sid_int,)
            )
            if existing:
                execute_stmt(conn, "DELETE FROM students WHERE id = %s", (sid_int,))
                deleted += 1
        if deleted:
            target = class_name or "multiple classes"
            log_change(
                conn,
                action="batch_delete_students",
                entity_type="students",
                class_name=class_name or None,
                details=f"Batch deleted {deleted} students from {target}.",
                affected_count=deleted,
            )
        conn.commit()
    except Exception as exc:
        conn.rollback()
        flash(f"Batch delete failed: {exc}", "danger")
        return redirect(url_for("manage_students", class_name=class_name or None, setup_tab=setup_tab))
    finally:
        conn.close()

    flash(f"✅ Deleted {deleted} student(s) and their marks.", "success")
    return redirect(url_for("manage_students", class_name=class_name or None, setup_tab=setup_tab))


@app.route("/clear_student_marks/<int:student_id>", methods=["POST"])
def clear_student_marks(student_id):
    if block_if_locked():
        return redirect(url_for("view_marks"))

    conn = get_db_connection()
    mark_count_row = fetch_one(
        conn,
        "SELECT COUNT(*) AS c FROM marks WHERE student_id = %s",
        (student_id,),
    )
    mark_count = mark_count_row["c"] if mark_count_row else 0
    student = fetch_one(
        conn,
        "SELECT roll_no, name, class_name FROM students WHERE id = %s",
        (student_id,),
    )
    execute_stmt(conn, "DELETE FROM marks WHERE student_id = %s", (student_id,))
    if mark_count:
        if student:
            details = (
                f"Cleared {mark_count} marks entries for {student['name']} "
                f"(roll {student['roll_no']}, {student['class_name']})."
            )
            class_name = student["class_name"]
        else:
            details = f"Cleared {mark_count} marks entries for student_id={student_id}."
            class_name = None
        log_change(
            conn,
            action="clear_student_marks",
            entity_type="marks",
            class_name=class_name,
            details=details,
            affected_count=mark_count,
        )
    conn.commit()
    conn.close()
    flash("Marks cleared.", "success")
    return redirect(url_for("view_marks"))


@app.route("/signature_image")
def signature_image():
    return send_from_directory(THIS_FOLDER, "sign.jpg")


@app.route("/admin/lock")
def admin_lock():
    return render_template("admin_lock.html")


@app.route("/admin/unlock", methods=["POST"])
def admin_unlock():
    pin = (request.form.get("master_pin") or "").strip()
    if not pin:
        flash("Enter Master PIN.", "warning")
        return redirect(url_for("admin_lock"))

    pin_hash = get_admin_pin_hash()
    if not pin_hash or not check_password_hash(pin_hash, pin):
        flash("Invalid Master PIN.", "danger")
        return redirect(url_for("admin_lock"))

    session["admin_unlocked"] = True
    conn = get_db_connection()
    try:
        log_change(
            conn,
            action="admin_unlock",
            entity_type="admin",
            details="Admin dashboard unlocked with Master PIN.",
            affected_count=1,
        )
        conn.commit()
    finally:
        conn.close()
    flash("Admin dashboard unlocked.", "success")
    return redirect(url_for("admin_dashboard"))


@app.route("/admin/logout", methods=["POST"])
def admin_logout():
    session.pop("admin_unlocked", None)
    flash("Admin dashboard locked.", "info")
    return redirect(url_for("admin_lock"))


@app.route("/admin/dashboard")
@admin_required
def admin_dashboard():
    settings = get_portal_settings()
    subjects_dict = get_subjects_dict()
    promotion_map = get_promotion_map()

    conn = get_db_connection()
    try:
        # Summary stats
        total_students_row = fetch_one(conn, "SELECT COUNT(*) AS c FROM students")
        total_marks_row = fetch_one(conn, "SELECT COUNT(*) AS c FROM marks")
        total_students = total_students_row["c"] if total_students_row else 0
        total_marks = total_marks_row["c"] if total_marks_row else 0

        # Per-class stats
        class_stats = []
        all_class_exams = get_all_class_exams()
        for cls, subjects in subjects_dict.items():
            sc_row = fetch_one(conn, "SELECT COUNT(*) AS c FROM students WHERE class_name = %s", (cls,))
            sc = sc_row["c"] if sc_row else 0
            mc_row = fetch_one(
                conn,
                "SELECT COUNT(*) AS c FROM marks m JOIN students s ON s.id=m.student_id WHERE s.class_name=%s",
                (cls,),
            )
            mc = mc_row["c"] if mc_row else 0
            exams = all_class_exams.get(cls, [])
            expected = sc * len(subjects) * max(len(exams), 1)
            pct = round(mc / expected * 100) if expected > 0 else 0
            class_stats.append({"class_name": cls, "students": sc, "marks": mc,
                                 "expected": expected, "pct": pct, "exams": len(exams)})

        # Recent activity (last 8)
        recent_logs = fetch_all(
            conn,
            "SELECT action, entity_type, details, created_at FROM change_logs ORDER BY id DESC LIMIT 8",
        )

        # DB size
        db_size_kb = 0
        if _is_sqlite_connection(conn) and os.path.exists(LOCAL_DB_FILE):
            db_size_kb = round(os.path.getsize(LOCAL_DB_FILE) / 1024, 1)
    finally:
        conn.close()

    return render_template(
        "admin_dashboard.html",
        settings=settings,
        all_classes=list(subjects_dict.keys()),
        promotion_map=promotion_map,
        total_students=total_students,
        total_marks=total_marks,
        class_stats=class_stats,
        recent_logs=recent_logs,
        db_size_kb=db_size_kb,
        homepage_panels=HOMEPAGE_PANELS,
    )


@app.route("/admin/promotion_map", methods=["POST"])
@admin_required
def admin_save_promotion_map():
    all_classes = list(get_subjects_dict().keys())
    new_map = {}
    for cls in all_classes:
        target = (request.form.get(f"promo_{cls}") or "").strip()
        if target:
            new_map[cls] = target
    conn = get_db_connection()
    try:
        set_setting_sqlite_safe(conn, "promotion_map_json", json.dumps(new_map))
        g._portal_settings = None  # invalidate cache
        log_change(conn, action="admin_update_promotion_map", entity_type="admin",
                   details="Updated promotion map.", affected_count=len(new_map))
        conn.commit()
    finally:
        conn.close()
    flash("Promotion map saved.", "success")
    return redirect(url_for("admin_dashboard"))


@app.route("/admin/settings", methods=["POST"])
@admin_required
def admin_save_settings():
    field_map = {
        "school_name": "School Name",
        "school_address": "School Address",
        "school_shortcode": "School Shortcode",
        "exam_name": "Examination Name",
        "academic_session": "Academic Session",
    }
    updates = {}
    for key in field_map:
        value = (request.form.get(key) or "").strip()
        if not value:
            flash(f"{field_map[key]} cannot be empty.", "warning")
            return redirect(url_for("admin_dashboard"))
        updates[key] = value

    conn = get_db_connection()
    try:
        for key, value in updates.items():
            set_setting_sqlite_safe(conn, key, value)
        log_change(
            conn,
            action="admin_update_settings",
            entity_type="admin",
            details=(
                "Updated portal settings: school name/address/shortcode, exam name, academic session."
            ),
            affected_count=len(updates),
        )
        conn.commit()
    finally:
        conn.close()
    g._portal_settings = None  # invalidate cache
    flash("Global settings saved. Site headers and printouts now use updated values.", "success")
    return redirect(url_for("admin_dashboard"))


@app.route("/admin/logo", methods=["POST"])
@admin_required
def admin_upload_logo():
    logo_file = request.files.get("school_logo")
    if not logo_file or not logo_file.filename:
        flash("Select a PNG logo file first.", "warning")
        return redirect(url_for("admin_dashboard"))

    raw_name = secure_filename(logo_file.filename)
    if not raw_name.lower().endswith(".png"):
        flash("Only PNG logo files are allowed.", "danger")
        return redirect(url_for("admin_dashboard"))

    os.makedirs(UPLOADS_FOLDER, exist_ok=True)
    logo_path = get_school_logo_path()
    logo_file.save(logo_path)

    # Resize to max 84×84px (2× retina for 42px display) to reduce bandwidth
    try:
        from PIL import Image as PilImage
        img = PilImage.open(logo_path)
        img.thumbnail((84, 84), PilImage.LANCZOS)
        img.save(logo_path, optimize=True)
    except Exception:
        pass  # Pillow unavailable or corrupt image — keep original

    updated_at = datetime.now().strftime("%Y%m%d%H%M%S")
    conn = get_db_connection()
    try:
        set_setting_sqlite_safe(conn, "school_logo_updated_at", updated_at)
        log_change(
            conn,
            action="admin_upload_logo",
            entity_type="admin",
            details=f"Uploaded school logo PNG ({raw_name}).",
            affected_count=1,
        )
        conn.commit()
    finally:
        conn.close()

    flash("School logo uploaded successfully.", "success")
    return redirect(url_for("admin_dashboard"))


@app.route("/admin/portal_lock", methods=["POST"])
@admin_required
def admin_toggle_portal_lock():
    lock_enabled = request.form.get("portal_locked") == "1"
    conn = get_db_connection()
    try:
        set_setting_sqlite_safe(conn, "portal_locked", "1" if lock_enabled else "0")
        log_change(
            conn,
            action="admin_portal_lock" if lock_enabled else "admin_portal_unlock",
            entity_type="admin",
            details=(
                "Portal editing locked; marks entry and student updates disabled."
                if lock_enabled
                else "Portal editing unlocked; marks entry and student updates enabled."
            ),
            affected_count=1,
        )
        conn.commit()
    finally:
        conn.close()

    flash("Portal has been locked." if lock_enabled else "Portal has been unlocked.", "success")
    return redirect(url_for("admin_dashboard"))


@app.route("/admin/export/master.xlsx")
@admin_required
def admin_export_master_xlsx():
    if Workbook is None:
        flash("Excel export dependency missing. Install openpyxl and retry.", "danger")
        return redirect(url_for("admin_dashboard"))

    conn = get_db_connection()
    try:
        settings = get_portal_settings(conn)
        students = fetch_all(
            conn,
            "SELECT id, class_name, roll_no, name, dob, father_name, mother_name FROM students ORDER BY class_name, CAST(roll_no AS UNSIGNED), roll_no",
        )
        marks = fetch_all(
            conn,
            """
            SELECT m.student_id, s.class_name, s.roll_no, s.name, m.subject, m.marks_obtained
            FROM marks m
            JOIN students s ON s.id = m.student_id
            ORDER BY s.class_name, CAST(s.roll_no AS UNSIGNED), s.roll_no, m.subject
            """,
        )
    finally:
        conn.close()

    workbook = Workbook()
    overview = workbook.active
    overview.title = "Portal Settings"
    overview.append(["Setting", "Value"])
    overview.append(["School Name", settings.get("school_name", "")])
    overview.append(["School Address", settings.get("school_address", "")])
    overview.append(["School Shortcode", settings.get("school_shortcode", "")])
    overview.append(["Examination Name", settings.get("exam_name", "")])
    overview.append(["Academic Session", settings.get("academic_session", "")])
    overview.append(["Portal Locked", "Yes" if setting_bool(settings.get("portal_locked", "0")) else "No"])
    overview.append(["Generated At", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])

    sheet_students = workbook.create_sheet("Students")
    sheet_students.append(["Student ID", "Class", "Roll No", "Name", "DOB", "Father Name", "Mother Name"])
    for row in students:
        sheet_students.append([row["id"], row["class_name"], row["roll_no"], row["name"], row["dob"], row["father_name"], row["mother_name"]])

    sheet_marks = workbook.create_sheet("Marks")
    sheet_marks.append(["Student ID", "Class", "Roll No", "Student Name", "Subject", "Marks"])
    for row in marks:
        sheet_marks.append(
            [
                row["student_id"],
                row["class_name"],
                row["roll_no"],
                row["name"],
                row["subject"],
                row["marks_obtained"],
            ]
        )

    summary_sheet = workbook.create_sheet("Class Summary")
    summary_sheet.append(["Class", "Students", "Appeared", "Pass", "Fail", "Absent", "Pass Rate (%)"])
    for class_name in get_subjects_dict().keys():
        results, _ = get_class_results(class_name)
        stats = calculate_class_stats(results)
        summary_sheet.append(
            [
                class_name,
                stats["total_students"],
                stats["appeared_count"],
                stats["pass_count"],
                stats["fail_count"],
                stats["absent_count"],
                stats["pass_rate"],
            ]
        )

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)

    conn = get_db_connection()
    try:
        log_change(
            conn,
            action="admin_export_master_xlsx",
            entity_type="admin",
            details="Downloaded master Excel export with settings, students, marks, and class summary.",
            affected_count=1,
        )
        conn.commit()
    finally:
        conn.close()

    filename = f"master_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return Response(
        output.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@app.route("/admin/backup/database")
@admin_required
def admin_backup_database():
    conn = get_db_connection()
    try:
        is_sqlite = _is_sqlite_connection(conn)
    finally:
        conn.close()

    if not is_sqlite:
        flash("Raw DB backup is available only in SQLite mode.", "warning")
        return redirect(url_for("admin_dashboard"))

    if not os.path.exists(LOCAL_DB_FILE):
        flash("Database file not found.", "danger")
        return redirect(url_for("admin_dashboard"))

    with open(LOCAL_DB_FILE, "rb") as f:
        payload = f.read()

    conn = get_db_connection()
    try:
        log_change(
            conn,
            action="admin_download_db_backup",
            entity_type="admin",
            details="Downloaded raw school_marks.db backup.",
            affected_count=1,
        )
        conn.commit()
    finally:
        conn.close()

    filename = f"school_marks_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.db"
    return Response(
        payload,
        mimetype="application/octet-stream",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@app.route("/admin/import/master_excel", methods=["POST"])
@admin_required
def admin_import_master_excel():
    if _load_workbook is None:
        flash("Excel dependency missing. Install openpyxl and retry.", "danger")
        return redirect(url_for("admin_dashboard"))

    excel_file = request.files.get("master_excel")
    if not excel_file or not excel_file.filename:
        flash("Select an Excel (.xlsx) file first.", "warning")
        return redirect(url_for("admin_dashboard"))

    if not (excel_file.filename or "").lower().endswith(".xlsx"):
        flash("Only .xlsx files are supported.", "warning")
        return redirect(url_for("admin_dashboard"))

    try:
        wb = _load_workbook(filename=io.BytesIO(excel_file.read()), read_only=True, data_only=True)
    except Exception as exc:
        flash(f"Failed to read Excel file: {exc}", "danger")
        return redirect(url_for("admin_dashboard"))

    def _build_col_map(ws):
        first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not first_row:
            return {}
        header = [str(c or "").strip().lower() for c in first_row]
        alias_groups = {
            "class_name": ["class", "class_name"],
            "roll_no": ["roll no", "roll_no", "roll", "roll number"],
            "name": ["name", "student name"],
            "dob": ["dob", "date of birth", "dateofbirth"],
            "father_name": ["father", "father_name", "father name"],
            "mother_name": ["mother", "mother_name", "mother name"],
            "subject": ["subject"],
            "marks_obtained": ["marks", "marks_obtained", "marks obtained"],
        }
        col_map = {}
        for key, aliases in alias_groups.items():
            for alias in aliases:
                if alias in header:
                    col_map[key] = header.index(alias)
                    break
        return col_map

    def _cell(row, col_map, key):
        idx = col_map.get(key)
        if idx is None or idx >= len(row):
            return ""
        val = row[idx]
        return str(val).strip() if val is not None else ""

    students_inserted = students_updated = students_unchanged = students_skipped = 0
    marks_inserted = marks_updated = marks_skipped = 0

    conn = get_db_connection()
    try:
        if "Students" in wb.sheetnames:
            ws = wb["Students"]
            col_map = _build_col_map(ws)
            if not {"class_name", "roll_no", "name"}.issubset(col_map.keys()):
                flash("Students sheet is missing required columns (Class, Roll No, Name). Students not imported.", "warning")
            else:
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not any(row):
                        continue
                    class_name = _cell(row, col_map, "class_name")
                    roll_no = _cell(row, col_map, "roll_no")
                    name = _cell(row, col_map, "name")
                    dob = _cell(row, col_map, "dob")
                    father_name = _cell(row, col_map, "father_name")
                    mother_name = _cell(row, col_map, "mother_name")
                    if not class_name or not roll_no or not name:
                        students_skipped += 1
                        continue
                    canonical_cls = canonicalize_class_name(class_name)
                    if not canonical_cls:
                        students_skipped += 1
                        continue
                    existing = fetch_one(
                        conn,
                        "SELECT id, name, dob, father_name, mother_name FROM students WHERE class_name = %s AND roll_no = %s",
                        (canonical_cls, roll_no),
                    )
                    if existing:
                        if (existing["name"] != name
                                or (existing["dob"] or "") != dob
                                or (existing["father_name"] or "") != father_name
                                or (existing["mother_name"] or "") != mother_name):
                            execute_stmt(
                                conn,
                                "UPDATE students SET name = %s, dob = %s, father_name = %s, mother_name = %s WHERE id = %s",
                                (name, dob, father_name, mother_name, existing["id"]),
                            )
                            students_updated += 1
                        else:
                            students_unchanged += 1
                    else:
                        _insert_sql = (
                            "INSERT OR IGNORE INTO students (roll_no, name, class_name, dob, father_name, mother_name) VALUES (%s, %s, %s, %s, %s, %s)"
                            if _is_sqlite_connection(conn)
                            else "INSERT IGNORE INTO students (roll_no, name, class_name, dob, father_name, mother_name) VALUES (%s, %s, %s, %s, %s, %s)"
                        )
                        execute_stmt(conn, _insert_sql, (roll_no, name, canonical_cls, dob, father_name, mother_name))
                        students_inserted += 1

        if "Marks" in wb.sheetnames:
            ws = wb["Marks"]
            col_map = _build_col_map(ws)
            if not {"class_name", "roll_no", "subject", "marks_obtained"}.issubset(col_map.keys()):
                flash("Marks sheet is missing required columns (Class, Roll No, Subject, Marks). Marks not imported.", "warning")
            else:
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not any(row):
                        continue
                    class_name = _cell(row, col_map, "class_name")
                    roll_no = _cell(row, col_map, "roll_no")
                    subject = _cell(row, col_map, "subject")
                    marks_raw = _cell(row, col_map, "marks_obtained")
                    if not class_name or not roll_no or not subject or not marks_raw:
                        marks_skipped += 1
                        continue
                    canonical_cls = canonicalize_class_name(class_name)
                    if not canonical_cls:
                        marks_skipped += 1
                        continue
                    try:
                        marks_val = int(float(marks_raw))
                    except (ValueError, TypeError):
                        marks_skipped += 1
                        continue
                    student = fetch_one(
                        conn,
                        "SELECT id FROM students WHERE class_name = %s AND roll_no = %s",
                        (canonical_cls, roll_no),
                    )
                    if not student:
                        marks_skipped += 1
                        continue
                    existing_mark = fetch_one(
                        conn,
                        "SELECT id, marks_obtained FROM marks WHERE student_id = %s AND subject = %s",
                        (student["id"], subject),
                    )
                    if existing_mark:
                        if existing_mark["marks_obtained"] != marks_val:
                            execute_stmt(
                                conn,
                                "UPDATE marks SET marks_obtained = %s WHERE id = %s",
                                (marks_val, existing_mark["id"]),
                            )
                            marks_updated += 1
                    else:
                        execute_stmt(
                            conn,
                            "INSERT INTO marks (student_id, subject, marks_obtained) VALUES (%s, %s, %s)",
                            (student["id"], subject, marks_val),
                        )
                        marks_inserted += 1

        log_change(
            conn,
            action="admin_import_master_excel",
            entity_type="admin",
            details=(
                f"Master Excel import: students {students_inserted} inserted, {students_updated} updated, "
                f"{students_unchanged} unchanged, {students_skipped} skipped; "
                f"marks {marks_inserted} inserted, {marks_updated} updated, {marks_skipped} skipped."
            ),
            affected_count=students_inserted + students_updated + marks_inserted + marks_updated,
        )
        conn.commit()
    except Exception as exc:
        conn.rollback()
        flash(f"Excel import failed: {exc}", "danger")
        return redirect(url_for("admin_dashboard"))
    finally:
        conn.close()

    flash(
        f"Excel import complete — students: {students_inserted} added, {students_updated} updated, "
        f"{students_unchanged} unchanged, {students_skipped} skipped; "
        f"marks: {marks_inserted} added, {marks_updated} updated, {marks_skipped} skipped.",
        "success",
    )
    return redirect(url_for("admin_dashboard"))


@app.route("/admin/reset/marks", methods=["POST"])
@admin_required
def admin_soft_reset_marks():
    confirmation = (request.form.get("confirm_phrase") or "").strip()
    if confirmation != "RESET MARKS":
        flash("Type RESET MARKS to confirm soft reset.", "warning")
        return redirect(url_for("admin_dashboard"))

    conn = get_db_connection()
    try:
        marks_count_row = fetch_one(conn, "SELECT COUNT(*) AS c FROM marks")
        marks_count = marks_count_row["c"] if marks_count_row else 0
        execute_stmt(conn, "DELETE FROM marks")
        execute_stmt(conn, "DELETE FROM change_logs WHERE entity_type = %s", ("marks",))
        log_change(
            conn,
            action="admin_soft_reset_marks",
            entity_type="admin",
            details=f"Soft reset executed. Deleted {marks_count} marks entries and related mark logs.",
            affected_count=marks_count,
        )
        conn.commit()
    finally:
        conn.close()

    flash("Soft reset complete: all marks and mark logs removed.", "success")
    return redirect(url_for("admin_dashboard"))


@app.route("/admin/reset/factory", methods=["POST"])
@admin_required
def admin_factory_reset():
    confirmation = (request.form.get("confirm_phrase") or "").strip()
    if confirmation != "FACTORY RESET":
        flash("Type FACTORY RESET to confirm factory reset.", "warning")
        return redirect(url_for("admin_dashboard"))

    conn = get_db_connection()
    try:
        counts_row = fetch_one(
            conn,
            """
            SELECT
                (SELECT COUNT(*) FROM students) AS students_count,
                (SELECT COUNT(*) FROM marks) AS marks_count
            """,
        )
        students_count = counts_row["students_count"] if counts_row else 0
        marks_count = counts_row["marks_count"] if counts_row else 0

        execute_stmt(conn, "DELETE FROM marks")
        execute_stmt(conn, "DELETE FROM students")
        execute_stmt(
            conn,
            "DELETE FROM change_logs WHERE entity_type IN (%s, %s)",
            ("marks", "students"),
        )
        log_change(
            conn,
            action="admin_factory_reset",
            entity_type="admin",
            details=(
                f"Factory reset executed. Deleted {students_count} students, {marks_count} marks, "
                "and related student/mark logs."
            ),
            affected_count=students_count + marks_count,
        )
        conn.commit()
    finally:
        conn.close()

    flash("Factory reset complete: students and marks wiped.", "success")
    return redirect(url_for("admin_dashboard"))


@app.route("/admin/reset/selective", methods=["POST"])
@admin_required
def admin_selective_reset():
    confirmation = (request.form.get("confirm_phrase") or "").strip()
    if confirmation != "SELECTIVE RESET":
        flash("Type SELECTIVE RESET exactly to confirm.", "warning")
        return redirect(url_for("admin_dashboard"))

    delete_marks = request.form.get("delete_marks") == "1"
    delete_students = request.form.get("delete_students") == "1"
    delete_logs = request.form.get("delete_logs") == "1"

    if not any([delete_marks, delete_students, delete_logs]):
        flash("Select at least one data type to delete.", "warning")
        return redirect(url_for("admin_dashboard"))

    conn = get_db_connection()
    try:
        parts = []
        affected = 0

        if delete_students:
            count_row = fetch_one(conn, "SELECT COUNT(*) AS c FROM students")
            count = count_row["c"] if count_row else 0
            execute_stmt(conn, "DELETE FROM students")  # CASCADE deletes marks
            parts.append(f"{count} students (and their marks)")
            affected += count
        elif delete_marks:
            count_row = fetch_one(conn, "SELECT COUNT(*) AS c FROM marks")
            count = count_row["c"] if count_row else 0
            execute_stmt(conn, "DELETE FROM marks")
            parts.append(f"{count} marks entries")
            affected += count

        if delete_logs:
            count_row = fetch_one(conn, "SELECT COUNT(*) AS c FROM change_logs")
            count = count_row["c"] if count_row else 0
            execute_stmt(conn, "DELETE FROM change_logs")
            parts.append(f"{count} activity log entries")
            affected += count

        if not delete_logs:
            log_change(
                conn,
                action="admin_selective_reset",
                entity_type="admin",
                details=f"Selective reset: deleted {', '.join(parts)}.",
                affected_count=affected,
            )

        conn.commit()
    except Exception as exc:
        conn.rollback()
        flash(f"Selective reset failed: {exc}", "danger")
        return redirect(url_for("admin_dashboard"))
    finally:
        conn.close()

    flash(f"Selective reset complete: deleted {', '.join(parts)}.", "success")
    return redirect(url_for("admin_dashboard"))


@app.route("/admin/change_pin", methods=["POST"])
@admin_required
def admin_change_pin():
    current_pin = (request.form.get("current_pin") or "").strip()
    new_pin = (request.form.get("new_pin") or "").strip()
    confirm_pin = (request.form.get("confirm_pin") or "").strip()

    if not current_pin or not new_pin or not confirm_pin:
        flash("All PIN fields are required.", "warning")
        return redirect(url_for("admin_dashboard") + "#tab-control")

    pin_hash = get_admin_pin_hash()
    if not pin_hash or not check_password_hash(pin_hash, current_pin):
        flash("Current PIN is incorrect.", "danger")
        return redirect(url_for("admin_dashboard") + "#tab-control")

    if new_pin != confirm_pin:
        flash("New PIN and confirmation do not match.", "danger")
        return redirect(url_for("admin_dashboard") + "#tab-control")

    if len(new_pin) < 4:
        flash("New PIN must be at least 4 characters.", "warning")
        return redirect(url_for("admin_dashboard") + "#tab-control")

    conn = get_db_connection()
    try:
        set_admin_pin_hash(conn, generate_password_hash(new_pin))
        log_change(conn, action="admin_change_pin", entity_type="admin",
                   details="Admin PIN changed successfully.", affected_count=1)
        conn.commit()
    finally:
        conn.close()

    flash("Admin PIN changed successfully.", "success")
    return redirect(url_for("admin_dashboard") + "#tab-control")


@app.route("/admin/reset/class_marks", methods=["POST"])
@admin_required
def admin_wipe_class_marks():
    class_name = (request.form.get("class_name") or "").strip()
    confirmation = (request.form.get("confirm_phrase") or "").strip()

    all_classes = list(get_subjects_dict().keys())
    if class_name not in all_classes:
        flash("Invalid class name.", "danger")
        return redirect(url_for("admin_dashboard") + "#tab-danger")

    if confirmation != f"WIPE {class_name}":
        flash(f'Type exactly: WIPE {class_name}', "warning")
        return redirect(url_for("admin_dashboard") + "#tab-danger")

    conn = get_db_connection()
    try:
        count_row = fetch_one(
            conn,
            "SELECT COUNT(*) AS c FROM marks m JOIN students s ON s.id = m.student_id WHERE s.class_name = %s",
            (class_name,),
        )
        count = count_row["c"] if count_row else 0
        execute_stmt(
            conn,
            "DELETE FROM marks WHERE student_id IN (SELECT id FROM students WHERE class_name = %s)",
            (class_name,),
        )
        log_change(conn, action="admin_wipe_class_marks", entity_type="admin",
                   class_name=class_name,
                   details=f"Wiped {count} marks for {class_name}.", affected_count=count)
        conn.commit()
    finally:
        conn.close()

    flash(f"Wiped {count} mark entries for {class_name}.", "success")
    return redirect(url_for("admin_dashboard") + "#tab-danger")


HOMEPAGE_PANELS = [
    ("enter_marks",    "Enter Marks",         "Main action card — subject-by-subject mark entry"),
    ("marks_table",    "Marks Table",         "Main action card — full class grid entry"),
    ("class_register", "Class-wise Register", "Main action card — view marks by class"),
    ("result_centre",  "Result Centre",       "Main action card — results, ledgers, report cards"),
    ("academic_setup", "Academic Setup",      "Secondary card — manage students/subjects/exams"),
    ("marks_progress", "Marks Progress",      "Secondary card — entry progress overview"),
    ("admin_panel",    "Admin Dashboard",     "Secondary card — admin shortcut link"),
    ("activity_logs",  "Activity Logs",       "Secondary card — change log"),
]


@app.route("/admin/panel_visibility", methods=["POST"])
@admin_required
def admin_save_panel_visibility():
    panel_ids = [p[0] for p in HOMEPAGE_PANELS]
    hidden = [pid for pid in panel_ids if request.form.get(f"hide_{pid}") == "1"]
    conn = get_db_connection()
    try:
        set_setting_sqlite_safe(conn, "hidden_panels", json.dumps(hidden))
        g._portal_settings = None
        log_change(conn, action="admin_panel_visibility", entity_type="admin",
                   details=f"Hidden panels updated: {hidden if hidden else 'none'}", affected_count=len(hidden))
        conn.commit()
    finally:
        conn.close()
    flash(f"Panel visibility saved. {len(hidden)} panel(s) hidden.", "success")
    return redirect(url_for("admin_dashboard") + "#tab-control")


@app.errorhandler(404)
def page_not_found(e):
    return render_template("404.html"), 404


@app.errorhandler(500)
def internal_error(e):
    return render_template("500.html"), 500


init_db()


if __name__ == "__main__":
    app.run(debug=False)
